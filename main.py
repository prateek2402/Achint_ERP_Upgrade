import os
import zipfile
import signal
import asyncio
import time
import re
import hashlib
import hmac
import base64
import secrets
import sqlite3
import shutil
import threading
import warnings
import io
from contextlib import asynccontextmanager
from pathlib import Path
from collections import defaultdict, deque
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Security, Request, Response, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from sqlalchemy import create_engine, func, or_
from sqlalchemy.orm import sessionmaker, Session, joinedload, selectinload
from pydantic import BaseModel

# google-generativeai is being phased out (its FutureWarning fires on import);
# we silence it locally and prefer google-genai when available. The wrapper
# below preserves the legacy `genai.GenerativeModel` interface so existing call
# sites + tests keep working. Both stacks are optional: minimal installs still
# run the ERP shell; extraction APIs require installing `requirements.txt`.
_legacy_genai = None
try:
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=FutureWarning, module=r"google\.generativeai.*")
        import google.generativeai as _legacy_genai
except Exception:
    pass
try:
    from google import genai as _new_genai  # google-genai package
    _HAS_NEW_GENAI = True
except Exception:
    _new_genai = None
    _HAS_NEW_GENAI = False

import jwt  # noqa: E402 — PyJWT (dist name `PyJWT`); must NOT install conflicting PyPI pkg `jwt`
if not callable(getattr(jwt, "encode", None)) or not callable(getattr(jwt, "decode", None)):
    raise ImportError(
        "The installed `jwt` module is not PyJWT (typically both PyPI packages `jwt` and `PyJWT` "
        "are installed). Uninstall the legacy one: pip uninstall jwt -y  "
        "then: pip install --force-reinstall 'PyJWT==2.12.1'"
    )
import datetime
import time
from typing import Optional, Any
import json
from decimal import Decimal, ROUND_HALF_UP

from app_logging import REQUEST_ID, configure_logging, get_logger, new_request_id

configure_logging()
log = get_logger("erp")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    Workbook = None
    Font = PatternFill = Alignment = None

from models import (
    Base, User, Client, PurchaseOrder, PoBaselineItem, 
    InvoiceDispatchItem, Invoice, PaymentHistory, 
    PaymentAllocation, SystemSettings, UnallocatedPaymentRegister, UnallocatedAdvanceRegister,
    AuditLog, UploadedDocument
)
from invoice_extract import parse_and_normalize_raw
from reconciliation import run_reconciliation

def load_local_env_file():
    env_path = ".env"
    if not os.path.exists(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if not key:
                    continue
                # Prefer .env values when the parent process exported an EMPTY var
                # (Python leaves key in environ with ''). Without this, GEMINI_API_KEY
                # in .env never applies and Gemini returns API_KEY_INVALID.
                prev_raw = os.environ.get(key)
                prev_empty = prev_raw is None or not str(prev_raw).strip()

                # API keys edited in `.env` for local dev must win when non-empty —
                # avoids stale system/user env Gemini keys masking the project's .env.
                if key == "GEMINI_API_KEY" and value:
                    os.environ[key] = value
                elif prev_empty:
                    os.environ[key] = value
    except Exception:
        # Keep startup resilient if .env is malformed.
        pass


load_local_env_file()

# --- Config & Setup ---
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()


def _normalize_contents_for_google_genai(contents: Any) -> Any:
    """google-generativeai accepted ``[prompt, {\"mime_type\": \"application/pdf\", \"data\": bytes}]``.
    google-genai validates ``contents`` as ``Part`` / ``Blob`` (``inline_data``), not snake_case blobs.
    """
    if not _HAS_NEW_GENAI or contents is None or not isinstance(contents, list):
        return contents

    from google.genai import types as genai_types

    normalized: list[Any] = []
    for item in contents:
        if isinstance(item, str):
            normalized.append(genai_types.Part(text=item))
            continue
        if isinstance(item, dict) and "data" in item:
            mime = item.get("mime_type") or item.get("mimeType") or "application/octet-stream"
            blob_bytes = item.get("data")
            if blob_bytes is None:
                blob_bytes = b""
            elif not isinstance(blob_bytes, (bytes, bytearray, memoryview)):
                blob_bytes = bytes(blob_bytes) if blob_bytes else b""
            else:
                blob_bytes = bytes(blob_bytes)

            normalized.append(
                genai_types.Part(inline_data=genai_types.Blob(mime_type=mime, data=blob_bytes))
            )
            continue

        normalized.append(item)

    return normalized


class _NewGenaiAdapter:
    """Adapt google-genai's Client.models.generate_content(...) to the legacy
    ``GenerativeModel(...).generate_content(parts, generation_config=...)`` API.

    Returns an object exposing ``.text`` so downstream code is unchanged.
    """

    def __init__(self, model_name: str):
        self.model_name = model_name
        self._client = _new_genai.Client(api_key=GEMINI_API_KEY) if GEMINI_API_KEY else None

    def generate_content(self, contents, generation_config=None, **_kwargs):
        if self._client is None:
            raise RuntimeError("Gemini API key not configured")
        cfg = None
        if generation_config and isinstance(generation_config, dict):
            mime = generation_config.get("response_mime_type")
            if mime:
                cfg = {"response_mime_type": mime}
        norm_contents = _normalize_contents_for_google_genai(contents)
        return self._client.models.generate_content(
            model=self.model_name,
            contents=norm_contents,
            config=cfg,
        )


class _GenAINamespace:
    """Compatibility namespace that mimics ``google.generativeai`` so existing
    callers (and test monkeypatches that target ``app_module.genai.GenerativeModel``)
    keep working. The runtime delegates to the new google-genai SDK when present
    and falls back to the legacy package otherwise.
    """

    def configure(self, api_key: str):  # noqa: D401 - mirror legacy signature
        if _HAS_NEW_GENAI:
            return
        if _legacy_genai is not None:
            _legacy_genai.configure(api_key=api_key)

    def GenerativeModel(self, name: str):  # noqa: N802 - keep legacy CamelCase
        if _HAS_NEW_GENAI:
            return _NewGenaiAdapter(name)
        if _legacy_genai is not None:
            return _legacy_genai.GenerativeModel(name)
        raise RuntimeError(
            "Gemini SDK is not installed. Install project requirements "
            "(google-genai / google-generativeai) to use invoice extraction features."
        )


genai = _GenAINamespace()
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

DATABASE_URL = os.getenv("APP_DATABASE_URL", "sqlite:///./erp_database.sqlite").strip()
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
if DATABASE_URL.startswith("sqlite:///"):
    sqlite_path = DATABASE_URL.replace("sqlite:///", "", 1)
    DB_FILE_PATH = Path(sqlite_path if sqlite_path else "erp_database.sqlite")
else:
    DB_FILE_PATH = Path("erp_database.sqlite")
BACKUP_DIR = Path(os.getenv("DB_BACKUP_DIR", "db_backups"))
BACKUP_INTERVAL_SECONDS = int(os.getenv("DB_BACKUP_INTERVAL_SECONDS", str(24 * 60 * 60)))
KEEP_BACKUPS_DAYS = int(os.getenv("KEEP_BACKUPS_DAYS", "30"))
_backup_thread_started = False
_write_serialization_lock = threading.Lock()

Base.metadata.create_all(bind=engine)


def _target_database_has_rows() -> bool:
    """Fail-safe guard for startup imports: any existing data means no auto-replace."""
    db = SessionLocal()
    try:
        for model in (Client, PurchaseOrder, Invoice, PaymentHistory, User, SystemSettings):
            if db.query(model.id).first() is not None:
                return True
        return False
    except Exception as exc:
        log.warning("legacy import skipped: could not verify target DB is empty: %s", exc)
        return True
    finally:
        db.close()


def _maybe_run_legacy_import():
    """Import legacy ERP snapshot on first boot when old_erp.sqlite is present."""
    legacy_path = Path(os.getenv("LEGACY_DB_PATH", "old_erp.sqlite"))
    marker = Path(".legacy_import_once.marker")
    if not legacy_path.exists():
        return
    if marker.exists():
        return
    if _target_database_has_rows():
        log.warning(
            "legacy import skipped: %s exists but target database already contains data",
            legacy_path,
        )
        return
    try:
        from migrate_sqlite import run_import

        run_import(force=False)
        log.info("legacy ERP data imported from %s", legacy_path)
    except Exception as exc:
        log.exception("legacy import failed: %s", exc)


def ensure_schema_columns():
    conn = sqlite3.connect(str(DB_FILE_PATH))
    cur = conn.cursor()
    try:
        cur.execute("PRAGMA table_info(clients)")
        client_cols = {row[1] for row in cur.fetchall()}
        if "display_currency" not in client_cols:
            cur.execute("ALTER TABLE clients ADD COLUMN display_currency TEXT DEFAULT 'INR'")
        if "exchange_rate" not in client_cols:
            cur.execute("ALTER TABLE clients ADD COLUMN exchange_rate REAL DEFAULT 83.0")

        cur.execute("PRAGMA table_info(purchase_orders)")
        cols = {row[1] for row in cur.fetchall()}
        if "contact_person" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN contact_person TEXT")
        if "project_name" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN project_name TEXT")
        if "is_completed" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN is_completed INTEGER DEFAULT 0")
        if "is_hidden" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN is_hidden INTEGER DEFAULT 0")
        if "completed_at" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN completed_at TEXT")
        if "tds_base" not in cols:
            cur.execute("ALTER TABLE purchase_orders ADD COLUMN tds_base TEXT DEFAULT 'basic'")

        cur.execute("PRAGMA table_info(po_baseline_items)")
        baseline_cols = {row[1] for row in cur.fetchall()}
        if "material_type" not in baseline_cols:
            cur.execute("ALTER TABLE po_baseline_items ADD COLUMN material_type TEXT")
        if "dispatch_alias" not in baseline_cols:
            cur.execute("ALTER TABLE po_baseline_items ADD COLUMN dispatch_alias TEXT")
        if "dispatch_rate" not in baseline_cols:
            cur.execute("ALTER TABLE po_baseline_items ADD COLUMN dispatch_rate REAL DEFAULT 0")

        cur.execute("PRAGMA table_info(invoice_dispatch_items)")
        dispatch_cols = {row[1] for row in cur.fetchall()}
        if "rate_per_uom" not in dispatch_cols:
            cur.execute("ALTER TABLE invoice_dispatch_items ADD COLUMN rate_per_uom REAL DEFAULT 0")

        cur.execute("PRAGMA table_info(system_settings)")
        settings_cols = {row[1] for row in cur.fetchall()}
        if "fy_start_month" not in settings_cols:
            cur.execute("ALTER TABLE system_settings ADD COLUMN fy_start_month INTEGER DEFAULT 4")
        if "fy_start_day" not in settings_cols:
            cur.execute("ALTER TABLE system_settings ADD COLUMN fy_start_day INTEGER DEFAULT 1")

        # Audit log table — created by Base.metadata.create_all on fresh installs;
        # this defensive ALTER path keeps existing DBs upgrade-safe.
        cur.execute(
            "CREATE TABLE IF NOT EXISTS audit_log ("
            " id INTEGER PRIMARY KEY AUTOINCREMENT,"
            " at_utc TIMESTAMP NOT NULL,"
            " user_id INTEGER,"
            " username TEXT,"
            " role TEXT,"
            " entity_type TEXT NOT NULL,"
            " entity_id TEXT,"
            " action TEXT NOT NULL,"
            " summary TEXT,"
            " details TEXT,"
            " ip_address TEXT)"
        )
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_at_utc ON audit_log(at_utc)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_entity_type ON audit_log(entity_type)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_entity_id ON audit_log(entity_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_action ON audit_log(action)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_user_id ON audit_log(user_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_audit_log_username ON audit_log(username)")

        # Uploaded document cache (idempotent invoice/PO PDF uploads).
        cur.execute(
            "CREATE TABLE IF NOT EXISTS uploaded_documents ("
            " id INTEGER PRIMARY KEY AUTOINCREMENT,"
            " sha256 VARCHAR(64) NOT NULL,"
            " kind VARCHAR(32) NOT NULL,"
            " original_filename TEXT,"
            " byte_size INTEGER,"
            " uploaded_by TEXT,"
            " uploaded_at TIMESTAMP NOT NULL,"
            " parsed_invoice_no TEXT,"
            " parsed_po_no TEXT,"
            " status VARCHAR(32) NOT NULL DEFAULT 'extracted',"
            " raw_data TEXT,"
            " parsed_json TEXT,"
            " warnings_json TEXT,"
            " parse_error TEXT)"
        )
        cur.execute("CREATE INDEX IF NOT EXISTS ix_uploaded_documents_sha256 ON uploaded_documents(sha256)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_uploaded_documents_kind ON uploaded_documents(kind)")

        conn.commit()
    finally:
        conn.close()


def prune_old_backups(keep_days: int) -> int:
    """Delete auto-generated backups older than ``keep_days``. Returns count removed.

    Only files matching ``erp_database_*.sqlite`` are touched; manual backups
    with other names are left alone, as is the restore-drill sandbox subdir.
    Setting KEEP_BACKUPS_DAYS=0 (or negative) disables rotation.
    """
    if keep_days <= 0 or not BACKUP_DIR.exists():
        return 0
    cutoff = time.time() - (keep_days * 86400)
    removed = 0
    for f in BACKUP_DIR.glob("erp_database_*.sqlite"):
        try:
            if f.is_file() and f.stat().st_mtime < cutoff:
                f.unlink()
                removed += 1
        except OSError as exc:
            log.warning("backup rotation: failed to delete %s: %s", f, exc)
    if removed:
        log.info("backup rotation removed %d file(s) older than %d day(s)", removed, keep_days)
    return removed


def perform_database_backup():
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"erp_database_{ts}.sqlite"
        if DB_FILE_PATH.exists():
            shutil.copy2(DB_FILE_PATH, backup_path)
            log.info("daily backup created: %s", backup_path)
        prune_old_backups(KEEP_BACKUPS_DAYS)
    except Exception as exc:
        log.warning("daily backup failed: %s", exc)


def start_daily_backup_worker():
    global _backup_thread_started
    if _backup_thread_started:
        return
    _backup_thread_started = True

    def _worker():
        while True:
            perform_database_backup()
            time.sleep(max(3600, BACKUP_INTERVAL_SECONDS))

    threading.Thread(target=_worker, name="daily-db-backup", daemon=True).start()


ensure_schema_columns()

APP_ENV = os.getenv("APP_ENV", "development").strip().lower()
ENABLE_API_DOCS = os.getenv("ENABLE_API_DOCS", "1" if APP_ENV != "production" else "0").strip() == "1"
APP_BOOT_TIME = time.time()


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Modern replacement for ``@app.on_event('startup')``.

    Runs schema bootstrap, kicks off the daily backup worker, and creates the
    optional bootstrap admin account if env-configured.
    """
    ensure_schema_columns()
    _maybe_run_legacy_import()
    start_daily_backup_worker()
    db = SessionLocal()
    try:
        bootstrap_username = os.getenv("BOOTSTRAP_ADMIN_USERNAME", "").strip()
        bootstrap_password = os.getenv("BOOTSTRAP_ADMIN_PASSWORD", "").strip()
        if bootstrap_username and bootstrap_password:
            existing = db.query(User).filter(User.username == bootstrap_username).first()
            if not existing:
                db.add(User(
                    username=bootstrap_username,
                    hashed_password=hash_password(bootstrap_password),
                    role="admin",
                ))
                db.commit()
                log.info("bootstrap admin account created for user: %s", bootstrap_username)
    finally:
        db.close()
    yield
    # No teardown required; daemon thread exits with the process.


app = FastAPI(
    title="Achint ERP API",
    docs_url="/docs" if ENABLE_API_DOCS else None,
    redoc_url="/redoc" if ENABLE_API_DOCS else None,
    openapi_url="/openapi.json" if ENABLE_API_DOCS else None,
    lifespan=lifespan,
)
SECRET_KEY = os.getenv("JWT_SECRET_KEY", "").strip()
if not SECRET_KEY:
    SECRET_KEY = secrets.token_urlsafe(48)
    log.warning("JWT_SECRET_KEY is not set; using ephemeral runtime secret. Set env var for stable secure authentication.")
JWT_ALGORITHM = "HS256"

PBKDF2_ROUNDS = int(os.getenv("PASSWORD_PBKDF2_ROUNDS", "210000"))
# How long an issued JWT stays valid. Default 24h is the sweet spot for an
# in-house ERP: long enough that finance staff don't get kicked out mid-day,
# short enough that a leaked token is not forever. Override with JWT_EXPIRY_HOURS.
JWT_EXPIRY_HOURS = max(1, int(os.getenv("JWT_EXPIRY_HOURS", "24")))
UPLOAD_RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("UPLOAD_RATE_LIMIT_WINDOW_SECONDS", "60"))
UPLOAD_RATE_LIMIT_MAX_REQUESTS = int(os.getenv("UPLOAD_RATE_LIMIT_MAX_REQUESTS", "10"))
MAX_UPLOAD_FILES_PER_REQUEST = int(os.getenv("MAX_UPLOAD_FILES_PER_REQUEST", "5"))
MAX_UPLOAD_FILE_SIZE_BYTES = int(os.getenv("MAX_UPLOAD_FILE_SIZE_BYTES", str(10 * 1024 * 1024)))
GEMINI_INVOICE_MODEL = os.getenv("GEMINI_INVOICE_MODEL", "gemini-2.5-flash").strip() or "gemini-2.5-flash"
GEMINI_PO_MODEL = os.getenv("GEMINI_PO_MODEL", "gemini-2.5-flash").strip() or "gemini-2.5-flash"
GEMINI_INVOICE_MAX_CONCURRENCY = max(1, int(os.getenv("GEMINI_INVOICE_MAX_CONCURRENCY", "3")))
GEMINI_INVOICE_JSON_RETRIES = max(0, int(os.getenv("GEMINI_INVOICE_JSON_RETRIES", "1")))
_upload_rate_limiter: dict[str, deque] = defaultdict(deque)
LOGIN_RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("LOGIN_RATE_LIMIT_WINDOW_SECONDS", "300"))
LOGIN_RATE_LIMIT_MAX_FAILURES = int(os.getenv("LOGIN_RATE_LIMIT_MAX_FAILURES", "10"))
_login_failures: dict[str, deque] = defaultdict(deque)

cors_origins_env = os.getenv("CORS_ALLOW_ORIGINS", "").strip()
if cors_origins_env:
    allow_origins = [o.strip() for o in cors_origins_env.split(",") if o.strip()]
else:
    allow_origins = [
        "http://localhost",
        "http://127.0.0.1",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)

# Compress > 1 KB responses (the dispatch grid + audit log payloads benefit a lot).
app.add_middleware(GZipMiddleware, minimum_size=1024)


@app.middleware("http")
async def add_request_id(request: Request, call_next):
    """Stamp every request + response with a short id for log correlation.

    Honours an inbound X-Request-Id when present so a chain of services / a
    reverse proxy can pre-assign one. Otherwise we generate a fresh one.
    """
    incoming = (request.headers.get("X-Request-Id") or "").strip()
    rid = incoming if incoming and len(incoming) <= 64 else new_request_id()
    token = REQUEST_ID.set(rid)
    try:
        response: Response = await call_next(request)
    finally:
        REQUEST_ID.reset(token)
    response.headers["X-Request-Id"] = rid
    return response


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, PBKDF2_ROUNDS)
    return f"pbkdf2_sha256${PBKDF2_ROUNDS}${base64.b64encode(salt).decode()}${base64.b64encode(digest).decode()}"


def verify_password(password: str, stored_password: str) -> bool:
    if not stored_password:
        return False
    if not stored_password.startswith("pbkdf2_sha256$"):
        return hmac.compare_digest(stored_password, password)
    try:
        _, rounds, salt_b64, digest_b64 = stored_password.split("$", 3)
        salt = base64.b64decode(salt_b64)
        expected = base64.b64decode(digest_b64)
        candidate = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(rounds))
        return hmac.compare_digest(candidate, expected)
    except Exception:
        return False


ALLOWED_USER_ROLES = {"admin", "logistics", "user"}
USERNAME_PATTERN = re.compile(r"^[A-Za-z0-9._-]{3,32}$")


def normalize_username(value: str) -> str:
    return (value or "").strip().lower()


def validate_password_strength(password: str):
    pwd = password or ""
    if len(pwd) < 10:
        raise HTTPException(status_code=400, detail="Password must be at least 10 characters long.")
    if not re.search(r"[A-Z]", pwd):
        raise HTTPException(status_code=400, detail="Password must include at least one uppercase letter.")
    if not re.search(r"[a-z]", pwd):
        raise HTTPException(status_code=400, detail="Password must include at least one lowercase letter.")
    if not re.search(r"\d", pwd):
        raise HTTPException(status_code=400, detail="Password must include at least one number.")
    if not re.search(r"[^A-Za-z0-9]", pwd):
        raise HTTPException(status_code=400, detail="Password must include at least one special character.")


def enforce_upload_rate_limit(user_key: str):
    now = time.time()
    q = _upload_rate_limiter[user_key]
    while q and now - q[0] > UPLOAD_RATE_LIMIT_WINDOW_SECONDS:
        q.popleft()
    if len(q) >= UPLOAD_RATE_LIMIT_MAX_REQUESTS:
        raise HTTPException(status_code=429, detail="Upload rate limit exceeded. Please retry later.")
    q.append(now)


def require_pdf_files(files: list[UploadFile]):
    if not files:
        raise HTTPException(status_code=400, detail="At least one PDF file is required.")
    if len(files) > MAX_UPLOAD_FILES_PER_REQUEST:
        raise HTTPException(status_code=400, detail=f"Maximum {MAX_UPLOAD_FILES_PER_REQUEST} files allowed per request.")
    for f in files:
        if (f.content_type or "").lower() not in {"application/pdf", "application/x-pdf"}:
            raise HTTPException(status_code=400, detail=f"Invalid content type for {f.filename}. Only PDF is allowed.")


def _client_ip_from_request(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for", "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def enforce_login_rate_limit(user_key: str):
    now = time.time()
    q = _login_failures[user_key]
    while q and now - q[0] > LOGIN_RATE_LIMIT_WINDOW_SECONDS:
        q.popleft()
    if len(q) >= LOGIN_RATE_LIMIT_MAX_FAILURES:
        raise HTTPException(status_code=429, detail="Too many failed login attempts. Please retry later.")


def record_login_failure(user_key: str):
    now = time.time()
    q = _login_failures[user_key]
    while q and now - q[0] > LOGIN_RATE_LIMIT_WINDOW_SECONDS:
        q.popleft()
    q.append(now)


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    write_methods = {"POST", "PUT", "PATCH", "DELETE"}
    if request.method.upper() in write_methods:
        with _write_serialization_lock:
            response: Response = await call_next(request)
    else:
        response: Response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store"
    if request.url.scheme == "https":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def compute_sha256(data: bytes) -> str:
    """Compute hex SHA-256 of raw bytes (single source of truth for upload dedupe)."""
    return hashlib.sha256(data or b"").hexdigest()


def find_cached_uploaded_document(db: Session, sha256_hex: str, kind: str) -> Optional[UploadedDocument]:
    """Return the most recent successful prior upload matching (hash, kind), if any."""
    if not sha256_hex or not kind:
        return None
    return (
        db.query(UploadedDocument)
        .filter(UploadedDocument.sha256 == sha256_hex, UploadedDocument.kind == kind)
        .order_by(UploadedDocument.id.desc())
        .first()
    )


def persist_uploaded_document(
    db: Session,
    *,
    sha256_hex: str,
    kind: str,
    filename: Optional[str],
    byte_size: int,
    uploaded_by: Optional[str],
    raw_data: Optional[str],
    parsed: Optional[dict],
    warnings: Optional[list],
    parse_error: Optional[str],
) -> None:
    """Best-effort cache write; failures must not abort the parent request."""
    try:
        parsed_invoice_no = None
        parsed_po_no = None
        if parsed and isinstance(parsed, dict):
            parsed_invoice_no = (parsed.get("invNo") or parsed.get("inv_no") or None) if kind == "invoice" else None
            parsed_po_no = (parsed.get("poNo") or parsed.get("po_no") or None)
        record = UploadedDocument(
            sha256=sha256_hex,
            kind=kind,
            original_filename=filename,
            byte_size=byte_size,
            uploaded_by=uploaded_by,
            uploaded_at=datetime.datetime.now(datetime.timezone.utc),
            parsed_invoice_no=parsed_invoice_no,
            parsed_po_no=parsed_po_no,
            status="extracted" if not parse_error else "parse_error",
            raw_data=raw_data,
            parsed_json=json.dumps(parsed, default=str, ensure_ascii=False) if parsed else None,
            warnings_json=json.dumps(warnings or [], ensure_ascii=False) if warnings else None,
            parse_error=parse_error,
        )
        db.add(record)
        db.commit()
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        log.warning("uploaded_documents cache write failed: %s", exc)


def record_audit(
    db: Session,
    user: Optional[User],
    entity_type: str,
    action: str,
    entity_id: Optional[Any] = None,
    summary: Optional[str] = None,
    details: Any = None,
    request: Optional[Request] = None,
    commit: bool = False,
) -> None:
    """Append-only audit trail writer.

    Best-effort: any exception here is swallowed so the underlying business
    transaction is never aborted by audit failures. ``details`` is JSON-encoded
    automatically when not already a string.
    """

    try:
        if isinstance(details, str) or details is None:
            details_text = details
        else:
            try:
                details_text = json.dumps(details, default=str, ensure_ascii=False)
            except Exception:
                details_text = str(details)

        ip = None
        if request is not None:
            try:
                ip = request.client.host if request.client else None
            except Exception:
                ip = None

        entry = AuditLog(
            at_utc=datetime.datetime.now(datetime.timezone.utc),
            user_id=getattr(user, "id", None),
            username=getattr(user, "username", None),
            role=getattr(user, "role", None),
            entity_type=str(entity_type)[:128],
            entity_id=None if entity_id is None else str(entity_id)[:128],
            action=str(action)[:32],
            summary=None if summary is None else str(summary)[:512],
            details=details_text,
            ip_address=ip,
        )
        db.add(entry)
        if commit:
            db.commit()
        else:
            db.flush()
    except Exception as exc:
        try:
            db.rollback()
        except Exception:
            pass
        log.warning("audit log write failed: %s", exc)

# --- Security Middleware ---
security_scheme = HTTPBearer()

def get_current_user(credentials: HTTPAuthorizationCredentials = Security(security_scheme), db: Session = Depends(get_db)):
    token = credentials.credentials
    try:
        payload = jwt.decode(
            token,
            SECRET_KEY,
            algorithms=[JWT_ALGORITHM],
            options={"require": ["exp", "id", "type"]},
        )
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id: int = payload.get("id")
        user = db.query(User).filter(User.id == user_id).first()
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# --- App Startup: handled by lifespan() above ---

# --- Pydantic Schemas ---
class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    password: str
    role: str

class PasswordChange(BaseModel):
    newPassword: str

class SelfPasswordChange(BaseModel):
    currentPassword: str
    newPassword: str

class ClientCreate(BaseModel):
    name: str

class ClientUpdate(BaseModel):
    name: str
    active: bool


class BaselineItemCreate(BaseModel):
    description: str
    ordered_qty: float
    inspected_qty: Optional[float] = 0.0 # NEW COLUMN
    uom: str
    material_type: Optional[str] = None
    dispatch_alias: Optional[str] = None
    dispatch_rate: Optional[float] = 0.0

class DispatchItemCreate(BaseModel):
    description: str
    qty: float
    uom: str
    inspected_qty: Optional[float] = 0.0
    rate_per_uom: Optional[float] = 0.0

class POCreate(BaseModel):
    client_id: int
    po_no: str
    contact_person: Optional[str] = None
    project_name: Optional[str] = None
    adv_pct: float = 0.0
    ret_pct: float = 0.0
    ret_base: str = "total"
    tds_base: str = "basic"
    tds_enabled: bool = False
    tds_rate: float = 0.0
    tds_threshold: float = 0.0
    baseline_items: list[BaselineItemCreate] = []

class InvoiceCreate(BaseModel):
    client_id: int
    po_no: Optional[str] = None
    invoice_no: str
    sub_entity: Optional[str] = None
    lr_no: Optional[str] = None
    inv_date: Optional[str] = None
    due_date: Optional[str] = None
    basic: float = 0.0
    gst: float = 0.0
    total: float = 0.0
    advance_adj: float = 0.0
    tds_ded: float = 0.0
    retention_held: float = 0.0
    net_payable: float = 0.0
    paid: float = 0.0
    balance: float = 0.0
    is_note: bool = False
    note_type: Optional[str] = None
    note_reason: Optional[str] = None
    dispatch_items: list[DispatchItemCreate] = [] # Ensure this matches the frontend key

class InvoiceUpdate(InvoiceCreate):
    pass


class LedgerExportRequest(BaseModel):
    invoice_nos: list[str]


class DispatchItemUpdate(BaseModel):
    description: str
    qty: float
    inspected_qty: Optional[float] = 0.0
    uom: Optional[str] = "Nos"
    merge_on_match: bool = True


class DispatchCellUpsert(BaseModel):
    invoice_no: str
    description: str
    uom: Optional[str] = "Nos"
    qty: float
    rate_per_uom: Optional[float] = 0.0


class DispatchColumnDeleteRequest(BaseModel):
    po_no: str
    description: str
    uom: Optional[str] = None
    client_id: Optional[int] = None


class DispatchColumnRenameRequest(BaseModel):
    po_no: str
    old_description: str
    old_uom: Optional[str] = None
    new_description: str
    new_uom: Optional[str] = None
    client_id: Optional[int] = None


class InvoiceInlineUpdate(BaseModel):
    inv_date: Optional[str] = None
    lr_no: Optional[str] = None
    total: Optional[float] = None

class PaymentUpdate(BaseModel):
    amount: float
    note: str

class PaymentAllocationTarget(BaseModel):
    inv_id: str
    amount: float

class PaymentAllocateRequest(BaseModel):
    client_id: int
    id: str
    date: str
    amount: float = 0.0
    note: Optional[str] = None
    mode: str = "cascade"  # cascade | targeted
    targets: list[PaymentAllocationTarget] = []
    hold_ret: bool = False
    hold_gst: bool = False
    only_gst: bool = False
    apply_adv: bool = False
    advance_only: bool = False
    fund_source: str = "receipt"  # receipt | unallocated
    move_to_po: Optional[str] = None
    po_no: Optional[str] = None
    source_po_nos: list[str] = []
    clear_po_pool: bool = False
    excess_action: str = "park"  # park | allocate_pending
    allow_overpayment: bool = False  # if True, allow allocating > invoice balance → negative balance

class TransferRequest(BaseModel):
    new_client_id: int
    action: str

class NoteIssueRequest(BaseModel):
    client_id: int
    note_no: str
    date: Optional[str] = None
    note_type: str
    amount: float
    reason: Optional[str] = None
    target_invoice_id: Optional[str] = None


class PoAdvanceManualApplyRequest(BaseModel):
    """Apply remaining PO advance pool to invoices per current PO terms (optional single invoice)."""
    po_no: str
    invoice_no: Optional[str] = None


# --- Auth & User Routes ---
@app.post("/api/login")
def login(req: LoginRequest, request: Request, db: Session = Depends(get_db)):
    normalized_username = normalize_username(req.username)
    user_key = f"{_client_ip_from_request(request)}:{normalized_username}"
    enforce_login_rate_limit(user_key)
    user = db.query(User).filter(func.lower(User.username) == normalized_username).first()
    if user and verify_password(req.password, user.hashed_password) and not user.hashed_password.startswith("pbkdf2_sha256$"):
        # Seamless upgrade path for legacy plaintext passwords.
        user.hashed_password = hash_password(req.password)
        db.commit()
    if not user:
        record_login_failure(user_key)
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if not verify_password(req.password, user.hashed_password):
        record_login_failure(user_key)
        raise HTTPException(status_code=401, detail="Invalid username or password")

    _login_failures.pop(user_key, None)
    now = datetime.datetime.now(datetime.timezone.utc)
    expires_at = now + datetime.timedelta(hours=JWT_EXPIRY_HOURS)
    payload = {
        "id": user.id,
        "username": user.username,
        "role": user.role,
        "type": "access",
        "iat": now,
        "nbf": now,
        "exp": expires_at,
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm=JWT_ALGORITHM)
    # Surface the expiry to the client so the SPA can transparently expire its
    # cached credentials at the same instant the server stops accepting them.
    return {
        "token": token,
        "username": user.username,
        "role": user.role,
        "expires_at": expires_at.isoformat(),
        "expires_in_seconds": int(JWT_EXPIRY_HOURS * 3600),
    }

@app.get("/api/users")
def get_users(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    users = db.query(User).all()
    return [{"id": u.id, "username": u.username, "role": u.role} for u in users]

@app.post("/api/users")
def create_user(user_data: UserCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    username = (user_data.username or "").strip()
    if not USERNAME_PATTERN.match(username):
        raise HTTPException(status_code=400, detail="Username must be 3-32 chars and use only letters, numbers, dot, underscore, or hyphen.")
    normalized_role = (user_data.role or "").strip().lower()
    if normalized_role not in ALLOWED_USER_ROLES:
        raise HTTPException(status_code=400, detail="Invalid role selected.")
    validate_password_strength(user_data.password)
    existing = db.query(User).filter(func.lower(User.username) == normalize_username(username)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists.")
    new_user = User(username=username, hashed_password=hash_password(user_data.password), role=normalized_role)
    db.add(new_user)
    db.commit()
    if inv.po_no and inv.po_no != "UNASSIGNED":
        po_obj = db.query(PurchaseOrder).filter(
            PurchaseOrder.client_id == inv.client_id,
            PurchaseOrder.po_no == inv.po_no
        ).first()
        if po_obj and float(po_obj.adv_pct or 0.0) > 0:
            added_by_payment: dict[str, float] = defaultdict(float)
            consumed_by_payment: dict[str, float] = defaultdict(float)
            po_allocs = db.query(PaymentAllocation).join(
                PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
            ).filter(
                PaymentHistory.client_id == inv.client_id,
                PaymentAllocation.target_po_no == inv.po_no,
                PaymentAllocation.alloc_type.in_(["po_advance", "po_advance_applied"]),
            ).all()
            for al in po_allocs:
                pid = str(al.payment_id or "").strip()
                if not pid:
                    continue
                if al.alloc_type == "po_advance":
                    added_by_payment[pid] += float(al.amount or 0.0)
                elif al.alloc_type == "po_advance_applied":
                    consumed_by_payment[pid] += float(al.amount or 0.0)

            base_amt = float(new_inv.basic or 0.0) if (po_obj.ret_base or "total") == "basic" else float(new_inv.total or 0.0)
            max_allowed = max(0.0, base_amt * (float(po_obj.adv_pct or 0.0) / 100.0))
            existing_applied = db.query(PaymentAllocation).filter(
                PaymentAllocation.alloc_type == "po_advance_applied",
                PaymentAllocation.target_po_no == inv.po_no,
                PaymentAllocation.target_inv_id == new_inv.invoice_no
            ).all()
            already_applied = sum(float(a.amount or 0.0) for a in existing_applied)
            shortfall = max(0.0, max_allowed - already_applied)

            if shortfall > 0:
                for pid, added_amt in added_by_payment.items():
                    remaining_amt = float(added_amt) - float(consumed_by_payment.get(pid, 0.0))
                    if remaining_amt <= 0:
                        continue
                    take = min(shortfall, remaining_amt)
                    if take <= 0:
                        continue
                    db.add(PaymentAllocation(
                        payment_id=pid,
                        alloc_type="po_advance_applied",
                        target_inv_id=new_inv.invoice_no,
                        target_po_no=inv.po_no,
                        note_id=None,
                        amount=float(take),
                    ))
                    shortfall -= take
                    if shortfall <= 0:
                        break
        db.commit()
    return {"success": True, "id": new_user.id}


@app.get("/api/users/me")
def get_my_account(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "role": current_user.role
    }


@app.post("/api/users/change-password")
def change_own_password(req: SelfPasswordChange, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not verify_password(req.currentPassword, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect.")
    validate_password_strength(req.newPassword)
    if verify_password(req.newPassword, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="New password must be different from current password.")
    current_user.hashed_password = hash_password(req.newPassword)
    db.commit()
    return {"success": True}


@app.put("/api/users/{user_id}/password")
def admin_reset_password(user_id: int, req: PasswordChange, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    user_to_update = db.query(User).filter(User.id == user_id).first()
    if not user_to_update:
        raise HTTPException(status_code=404, detail="User not found.")
    validate_password_strength(req.newPassword)
    user_to_update.hashed_password = hash_password(req.newPassword)
    db.commit()
    return {"success": True}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if current_user.id == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete own account.")
    user_to_delete = db.query(User).filter(User.id == user_id).first()
    if user_to_delete:
        db.delete(user_to_delete)
        db.commit()
    return {"success": True}

class SettingsSchema(BaseModel):
    exchangeRate: float
    customColumns: list


class FiscalYearConfigSchema(BaseModel):
    startMonth: int
    startDay: int


class ClientCurrencySchema(BaseModel):
    displayCurrency: str
    exchangeRate: float


class POStatusSchema(BaseModel):
    is_completed: bool
    is_hidden: bool = False

# --- PHASE 4: GLOBAL SETTINGS ---
@app.get("/api/settings")
def get_settings(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access financial settings.")
    settings = db.query(SystemSettings).first()
    if not settings:
        settings = SystemSettings(exchange_rate=83.0, custom_columns="[]")
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return {
        "exchangeRate": settings.exchange_rate,
        "customColumns": json.loads(settings.custom_columns),
        "fiscalYearStart": {"month": int(settings.fy_start_month or 4), "day": int(settings.fy_start_day or 1)}
    }

@app.post("/api/settings")
def update_settings(data: SettingsSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    settings = db.query(SystemSettings).first()
    if not settings:
        settings = SystemSettings()
        db.add(settings)
        
    settings.exchange_rate = data.exchangeRate
    settings.custom_columns = json.dumps(data.customColumns)
    db.commit()
    return {"success": True}


@app.put("/api/settings/fiscal-year")
def update_fiscal_year_settings(data: FiscalYearConfigSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if data.startMonth < 1 or data.startMonth > 12 or data.startDay < 1 or data.startDay > 31:
        raise HTTPException(status_code=400, detail="Invalid fiscal year start date.")
    settings = db.query(SystemSettings).first()
    if not settings:
        settings = SystemSettings()
        db.add(settings)
    settings.fy_start_month = data.startMonth
    settings.fy_start_day = data.startDay
    db.commit()
    return {"success": True}


@app.put("/api/clients/{client_id}/currency")
def update_client_currency(client_id: int, data: ClientCurrencySchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    currency = (data.displayCurrency or "INR").upper().strip()
    if currency not in ("INR", "USD"):
        raise HTTPException(status_code=400, detail="Currency must be INR or USD.")
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    client.display_currency = currency
    client.exchange_rate = max(0.0001, float(data.exchangeRate or 83.0))
    db.commit()
    return {"success": True}


# --- PHASE 3: STRICT RELATIONAL CLIENT ENDPOINTS ---
@app.get("/api/clients")
def get_clients(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access client financial records.")
    clients = db.query(Client).all()
    return [{
        "id": c.id,
        "name": c.name,
        "active": c.active,
        "excess_funds": c.excess_funds,
        "display_currency": c.display_currency or "INR",
        "exchange_rate": c.exchange_rate or 83.0
    } for c in clients]

@app.post("/api/clients")
def create_client(client: ClientCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    existing = db.query(Client).filter(Client.name == client.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Account already exists.")
    new_client = Client(name=client.name, active=True, excess_funds=0.0, display_currency="INR", exchange_rate=83.0)
    db.add(new_client)
    db.commit()
    return {"success": True, "id": new_client.id, "name": new_client.name, "display_currency": "INR", "exchange_rate": 83.0}

@app.delete("/api/clients/{client_id}")
def delete_client(client_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    client = db.query(Client).filter(Client.id == client_id).first()
    if client:
        db.delete(client)
        db.commit()
    return {"success": True}

@app.get("/api/purchase-orders")
def get_purchase_orders(include_completed: bool = True, include_hidden: bool = False, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access PO financial configuration.")
    query = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items))
    if not include_completed:
        query = query.filter(PurchaseOrder.is_completed == False)
    if not include_hidden:
        query = query.filter(PurchaseOrder.is_hidden == False)
    pos = query.all()
    result = []
    for po in pos:
        items = [{
            "description": item.description,
            "ordered_qty": item.ordered_qty,
            "inspected_qty": item.inspected_qty,
            "uom": item.uom,
            "material_type": item.material_type,
            "dispatch_alias": item.dispatch_alias,
            "dispatch_rate": float(item.dispatch_rate or 0.0)
        } for item in po.baseline_items]
        result.append({
            "id": po.id,
            "client_id": po.client_id,
            "po_no": po.po_no,
            "contact_person": po.contact_person,
            "project_name": po.project_name,
            "is_completed": bool(po.is_completed),
            "is_hidden": bool(po.is_hidden),
            "completed_at": po.completed_at.isoformat() if po.completed_at else None,
            "adv_pct": po.adv_pct,
            "ret_pct": po.ret_pct,
            "ret_base": po.ret_base,
            "tds_base": getattr(po, "tds_base", None) or "basic",
            "tds_enabled": po.tds_enabled,
            "tds_rate": po.tds_rate,
            "tds_threshold": po.tds_threshold,
            "baseline_items": items
        })
    return result

@app.post("/api/purchase-orders")
def create_purchase_order(po: POCreate, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    
    existing_po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po.po_no).first()
    
    if existing_po:
        # UPSERT: Update the existing lazily-created PO with strict financial terms
        existing_po.contact_person = po.contact_person
        existing_po.project_name = po.project_name
        if current_user.role == "admin":
            existing_po.adv_pct = po.adv_pct
            existing_po.ret_pct = po.ret_pct
            existing_po.ret_base = po.ret_base
            existing_po.tds_base = po.tds_base
            existing_po.tds_enabled = po.tds_enabled
            existing_po.tds_rate = po.tds_rate
            existing_po.tds_threshold = po.tds_threshold
        po_id = existing_po.id
    else:
        # INSERT: Brand new PO from the terms configurator
        new_po = PurchaseOrder(
            client_id=po.client_id, po_no=po.po_no, contact_person=po.contact_person, project_name=po.project_name,
            adv_pct=(po.adv_pct if current_user.role == "admin" else 0.0),
            ret_pct=(po.ret_pct if current_user.role == "admin" else 0.0),
            ret_base=(po.ret_base if current_user.role == "admin" else "total"),
            tds_base=(po.tds_base if current_user.role == "admin" else "basic"),
            tds_enabled=(po.tds_enabled if current_user.role == "admin" else False),
            tds_rate=(po.tds_rate if current_user.role == "admin" else 0.0),
            tds_threshold=(po.tds_threshold if current_user.role == "admin" else 0.0)
        )
        db.add(new_po)
        db.flush()
        po_id = new_po.id
    
    # Safely clear and rewrite digital SKU baseline items
    db.query(PoBaselineItem).filter(PoBaselineItem.po_id == po_id).delete()
    
    for item in po.baseline_items:
        mt = (item.material_type or "").strip().lower() if item.material_type else None
        if mt not in ("brick", "castable_mortar"):
            mt = None
        new_item = PoBaselineItem(
            po_id=po_id,
            description=item.description,
            ordered_qty=item.ordered_qty,
            inspected_qty=item.inspected_qty,
            uom=item.uom,
            material_type=mt,
            dispatch_alias=(item.dispatch_alias or "").strip() or None,
            dispatch_rate=max(0.0, float(item.dispatch_rate or 0.0))
        )
        db.add(new_item)

    record_audit(
        db,
        current_user,
        entity_type="purchase_order",
        entity_id=po.po_no,
        action=("update" if existing_po else "create"),
        summary=f"{'Updated' if existing_po else 'Created'} PO {po.po_no} for client {po.client_id}",
        details={
            "client_id": po.client_id,
            "po_no": po.po_no,
            "contact_person": po.contact_person,
            "project_name": po.project_name,
            "adv_pct": po.adv_pct,
            "ret_pct": po.ret_pct,
            "tds_enabled": po.tds_enabled,
            "tds_rate": po.tds_rate,
            "baseline_item_count": len(po.baseline_items or []),
        },
        request=request,
    )
    db.commit()
    _auto_apply_po_advance(po.client_id, db, po.po_no)
    # Terms changes (adv/ret/tds config) must immediately reflect in ledger math views.
    recalculate_client_ledger(po.client_id, db)
    return {"success": True, "po_id": po_id}


@app.put("/api/purchase-orders/{po_no:path}/status")
def update_purchase_order_status(po_no: str, status: POStatusSchema, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    po_query = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no)
    if payload.client_id:
        po_query = po_query.filter(PurchaseOrder.client_id == payload.client_id)
    po = po_query.first()
    if not po:
        raise HTTPException(status_code=404, detail="PO not found")
    po.is_completed = bool(status.is_completed)
    po.is_hidden = bool(status.is_hidden)
    po.completed_at = datetime.date.today() if po.is_completed else None
    db.commit()
    return {"success": True}

@app.delete("/api/purchase-orders/{po_no:path}")
def delete_purchase_order(po_no: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    po_query = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no)
    if payload.client_id:
        po_query = po_query.filter(PurchaseOrder.client_id == payload.client_id)
    po = po_query.first()
    if po:
        client_id = po.client_id
        db.delete(po)
        db.commit()
        recalculate_client_ledger(client_id, db)
    return {"success": True}


@app.get("/api/logistics/dispatch-summary")
def get_logistics_dispatch_summary(
    status: Optional[str] = Query(default="", description="pending | cleared | empty"),
    search: Optional[str] = Query(default="", description="Search by client name or PO number"),
    sort_by: Optional[str] = Query(default="pending_qty", description="pending_qty | completion | client | po_no"),
    sort_dir: Optional[str] = Query(default="desc", description="asc | desc"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role not in ["admin", "logistics"]:
        raise HTTPException(status_code=403, detail="Admin or Logistics access required")

    def _empty_dispatch_summary():
        return {
            "overview": {
                "total_statements": 0,
                "pending_statements": 0,
                "total_pending_qty": 0.0,
                "total_ordered_qty": 0.0,
                "total_dispatched_qty": 0.0,
                "completion_pct": 0.0
            },
            "rows": []
        }

    active_clients = db.query(Client).filter(or_(Client.active == True, Client.active.is_(None))).all()
    if not active_clients:
        return _empty_dispatch_summary()

    client_by_id = {c.id: c.name for c in active_clients}
    client_ids = list(client_by_id.keys())

    pos = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).filter(
        PurchaseOrder.client_id.in_(client_ids)
    ).all()

    po_by_id = {po.id: po for po in pos}
    po_ids = list(po_by_id.keys())
    if not po_ids:
        return _empty_dispatch_summary()

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id.in_(po_ids),
        Invoice.is_note == False
    ).all()

    dispatch_map: dict[tuple[int, str], float] = defaultdict(float)
    dispatch_inspected_map: dict[tuple[int, str], float] = defaultdict(float)
    po_dispatch_totals: dict[int, float] = defaultdict(float)
    po_dispatch_descriptions: dict[int, set[str]] = defaultdict(set)
    for inv in invoices:
        if not inv.po_id:
            continue
        for item in inv.dispatch_items:
            key = (inv.po_id, (item.description or "").strip().upper())
            if key[1]:
                dispatch_map[key] += float(item.dispatched_qty or 0.0)
                dispatch_inspected_map[key] += float(item.inspected_qty or 0.0)
                po_dispatch_totals[inv.po_id] += float(item.dispatched_qty or 0.0)
                po_dispatch_descriptions[inv.po_id].add(key[1])

    rows = []
    total_ordered_qty = 0.0
    total_dispatched_qty = 0.0
    total_pending_qty = 0.0
    pending_statements = 0

    for po in pos:
        baseline_items = po.baseline_items or []
        if not baseline_items:
            # Legacy-safe fallback: show statement rows even when baseline lines are missing.
            dispatched_fallback = float(po_dispatch_totals.get(po.id, 0.0))
            if dispatched_fallback <= 0.0001:
                continue
            rows.append({
                "client": client_by_id.get(po.client_id, "Unknown"),
                "po_no": po.po_no,
                "project_name": po.project_name,
                "contact_person": po.contact_person,
                "material_lines": len(po_dispatch_descriptions.get(po.id, set())),
                "pending_lines": 0,
                "ordered_qty": dispatched_fallback,
                "dispatched_qty": dispatched_fallback,
                "inspected_qty": 0.0,
                "pending_qty": 0.0,
                "completion": 100.0
            })
            total_ordered_qty += dispatched_fallback
            total_dispatched_qty += dispatched_fallback
            continue

        ordered_qty = 0.0
        dispatched_qty = 0.0
        inspected_qty = 0.0
        pending_qty = 0.0
        pending_lines = 0

        for base in baseline_items:
            ordered = float(base.ordered_qty or 0.0)
            inspected = float(base.inspected_qty or 0.0)
            desc_key = (base.description or "").strip().upper()
            dispatched = dispatch_map.get((po.id, desc_key), 0.0) if desc_key else 0.0
            inspected_actual = dispatch_inspected_map.get((po.id, desc_key), 0.0) if desc_key else 0.0
            pending = max(0.0, ordered - dispatched)

            ordered_qty += ordered
            dispatched_qty += dispatched
            inspected_qty += max(inspected, inspected_actual)
            pending_qty += pending
            if pending > 0.0001:
                pending_lines += 1

        completion = min(100.0, (dispatched_qty / ordered_qty) * 100.0) if ordered_qty > 0 else 0.0
        if pending_qty > 0.0001:
            pending_statements += 1

        total_ordered_qty += ordered_qty
        total_dispatched_qty += dispatched_qty
        total_pending_qty += pending_qty

        rows.append({
            "client": client_by_id.get(po.client_id, "Unknown"),
            "po_no": po.po_no,
            "project_name": po.project_name,
            "contact_person": po.contact_person,
            "material_lines": len(baseline_items),
            "pending_lines": pending_lines,
            "ordered_qty": ordered_qty,
            "dispatched_qty": dispatched_qty,
            "inspected_qty": inspected_qty,
            "pending_qty": pending_qty,
            "completion": completion
        })

    status_val = (status or "").strip().lower()
    search_val = (search or "").strip().lower()

    filtered_rows = []
    for row in rows:
        if status_val == "pending" and row["pending_qty"] <= 0.0001:
            continue
        if status_val == "cleared" and row["pending_qty"] > 0.0001:
            continue
        if search_val and search_val not in row["client"].lower() and search_val not in row["po_no"].lower():
            continue
        filtered_rows.append(row)

    sort_key = (sort_by or "pending_qty").strip().lower()
    reverse_sort = (sort_dir or "desc").strip().lower() != "asc"
    if sort_key == "completion":
        filtered_rows.sort(key=lambda x: x["completion"], reverse=reverse_sort)
    elif sort_key == "client":
        filtered_rows.sort(key=lambda x: x["client"].lower(), reverse=reverse_sort)
    elif sort_key == "po_no":
        filtered_rows.sort(key=lambda x: x["po_no"].lower(), reverse=reverse_sort)
    else:
        filtered_rows.sort(key=lambda x: x["pending_qty"], reverse=reverse_sort)

    ro = round_qty_total(total_ordered_qty)
    rd = round_qty_total(total_dispatched_qty)
    rp = round_qty_total(total_pending_qty)
    completion_pct = round((min(rd, ro) / ro * 100.0) if ro > 0 else 0.0, 2)
    return {
        "overview": {
            "total_statements": len(rows),
            "pending_statements": pending_statements,
            "total_pending_qty": rp,
            "total_ordered_qty": ro,
            "total_dispatched_qty": rd,
            "completion_pct": completion_pct
        },
        "rows": filtered_rows
    }


@app.get("/api/logistics/dispatch-detail")
def get_logistics_dispatch_detail(
    client: str = Query(..., description="Client name"),
    po_no: str = Query(..., description="Purchase order number"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role not in ["admin", "logistics"]:
        raise HTTPException(status_code=403, detail="Admin or Logistics access required")

    client_name = (client or "").strip()
    po_no_val = (po_no or "").strip()
    if not client_name or not po_no_val:
        raise HTTPException(status_code=400, detail="client and po_no are required")

    po = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).join(Client).filter(
        Client.name == client_name,
        Client.active == True,
        PurchaseOrder.po_no == po_no_val
    ).first()
    if not po:
        raise HTTPException(status_code=404, detail="Dispatch statement not found")

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id == po.id,
        Invoice.is_note == False
    ).all()

    dispatch_by_desc: dict[str, float] = defaultdict(float)
    invoice_rows = []
    for inv in invoices:
        for item in inv.dispatch_items:
            key = (item.description or "").strip().upper()
            qty = float(item.dispatched_qty or 0.0)
            if key:
                dispatch_by_desc[key] += qty
            invoice_rows.append({
                "dispatch_item_id": item.id,
                "invoice_no": inv.invoice_no,
                "invoice_date": inv.inv_date.isoformat() if inv.inv_date else None,
                "description": item.description or "",
                "dispatched_qty": qty,
                "inspected_qty": float(item.inspected_qty or 0.0),
                "uom": item.uom or ""
            })

    summary_rows = []
    total_ordered = 0.0
    total_dispatched = 0.0
    for base in (po.baseline_items or []):
        ordered = float(base.ordered_qty or 0.0)
        dispatched = float(dispatch_by_desc.get((base.description or "").strip().upper(), 0.0))
        pending = max(0.0, ordered - dispatched)
        total_ordered += ordered
        total_dispatched += dispatched
        summary_rows.append({
            "description": base.description or "",
            "uom": base.uom or "",
            "ordered_qty": ordered,
            "inspected_qty": float(base.inspected_qty or 0.0),
            "material_type": base.material_type or None,
            "dispatch_alias": base.dispatch_alias or None,
            "dispatch_rate": float(base.dispatch_rate or 0.0),
            "dispatched_qty": dispatched,
            "pending_qty": pending
        })

    invoice_rows.sort(
        key=lambda r: (*invoice_ledger_sort_key_from_strings(r.get("invoice_date"), r.get("invoice_no")), r.get("dispatch_item_id") or "")
    )
    ro = round_qty_total(total_ordered)
    rd = round_qty_total(total_dispatched)
    rp = round_qty_total(max(0.0, total_ordered - total_dispatched))
    completion = round(min(100.0, (rd / ro) * 100.0) if ro > 0 else 0.0, 2)
    return {
        "overview": {
            "client": client_name,
            "po_no": po_no_val,
            "project_name": po.project_name,
            "contact_person": po.contact_person,
            "ordered_qty": ro,
            "dispatched_qty": rd,
            "pending_qty": rp,
            "completion_pct": completion
        },
        "summary_rows": summary_rows,
        "invoice_rows": invoice_rows
    }


def _fill_dispatch_invoice_sheet(
    ws_invoice,
    *,
    client_name: str,
    po_no_val: str,
    invoice_rows: list,
    title_font: Font,
    header_fill: PatternFill,
    header_font: Font,
) -> None:
    ws_invoice["A1"] = f"Invoice-wise Dispatch: {client_name} | PO: {po_no_val}"
    ws_invoice["A1"].font = title_font
    ws_invoice.merge_cells("A1:F1")
    ws_invoice.append(["Invoice No", "Invoice Date", "Material Description", "Dispatched Qty", "Inspected Qty", "UOM"])
    for c in ws_invoice[2]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
    for row in invoice_rows:
        ws_invoice.append([
            row["invoice_no"],
            row["invoice_date"],
            row["description"],
            round(row["dispatched_qty"], 2),
            round(row["inspected_qty"], 2),
            row["uom"],
        ])


@app.get("/api/dispatch/export-xlsx")
def export_dispatch_detail_xlsx(
    client: str = Query(..., description="Client name"),
    po_no: str = Query(..., description="Purchase order number"),
    layout: str = Query(
        "both",
        description="Excel layout: both (default), consolidated (item summary only), invoice_wise (per-invoice lines only)",
    ),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Excel export dependency missing. Install openpyxl.")
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")

    client_name = (client or "").strip()
    po_no_val = (po_no or "").strip()
    if not client_name or not po_no_val:
        raise HTTPException(status_code=400, detail="client and po_no are required")

    po = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).join(Client).filter(
        Client.name == client_name,
        PurchaseOrder.po_no == po_no_val
    ).first()
    if not po:
        raise HTTPException(status_code=404, detail="Dispatch statement not found")

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id == po.id,
        Invoice.is_note == False
    ).all()

    dispatch_by_desc: dict[str, float] = defaultdict(float)
    invoice_rows = []
    for inv in invoices:
        for item in inv.dispatch_items:
            key = (item.description or "").strip().upper()
            qty = float(item.dispatched_qty or 0.0)
            if key:
                dispatch_by_desc[key] += qty
            invoice_rows.append({
                "invoice_no": inv.invoice_no,
                "invoice_date": inv.inv_date.isoformat() if inv.inv_date else "",
                "description": item.description or "",
                "dispatched_qty": qty,
                "inspected_qty": float(item.inspected_qty or 0.0),
                "uom": item.uom or ""
            })

    summary_rows = []
    total_ordered = 0.0
    total_dispatched = 0.0
    for base in (po.baseline_items or []):
        ordered = float(base.ordered_qty or 0.0)
        dispatched = float(dispatch_by_desc.get((base.description or "").strip().upper(), 0.0))
        pending = max(0.0, ordered - dispatched)
        total_ordered += ordered
        total_dispatched += dispatched
        summary_rows.append({
            "description": base.description or "",
            "uom": base.uom or "",
            "ordered_qty": ordered,
            "inspected_qty": float(base.inspected_qty or 0.0),
            "material_type": base.material_type or None,
            "dispatch_alias": base.dispatch_alias or None,
            "dispatch_rate": float(base.dispatch_rate or 0.0),
            "dispatched_qty": dispatched,
            "pending_qty": pending
        })

    invoice_rows.sort(
        key=lambda r: (*invoice_ledger_sort_key_from_strings(r.get("invoice_date"), r.get("invoice_no")), r.get("description") or "")
    )

    layout_key = re.sub(r"[\s-]+", "_", (layout or "both").strip().lower())
    if layout_key in ("", "both", "all"):
        want_summary, want_invoice = True, True
    elif layout_key in ("consolidated", "summary", "item", "item_wise"):
        want_summary, want_invoice = True, False
    elif layout_key in ("invoice_wise", "invoice", "invoices", "per_invoice"):
        want_summary, want_invoice = False, True
    else:
        raise HTTPException(
            status_code=400,
            detail="layout must be 'both', 'consolidated', or 'invoice_wise'",
        )

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=12)

    wb = Workbook()
    created_sheets = []

    if want_summary:
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = f"Dispatch Statement: {client_name} | PO: {po_no_val}"
        ws_summary["A1"].font = title_font
        ws_summary.merge_cells("A1:F1")
        ws_summary.append(["Material Description", "UOM", "Ordered Qty", "Dispatched Qty", "Inspected Qty", "Pending Qty"])
        for c in ws_summary[2]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        for row in summary_rows:
            ws_summary.append([
                row["description"],
                row["uom"],
                round(row["ordered_qty"], 2),
                round(row["dispatched_qty"], 2),
                round(row["inspected_qty"], 2),
                round(row["pending_qty"], 2),
            ])
        pending_total = round_qty_total(max(0.0, total_ordered - total_dispatched))
        ws_summary.append([
            "TOTAL", "",
            round_qty_total(total_ordered),
            round_qty_total(total_dispatched),
            "",
            pending_total,
        ])
        for c in ws_summary[ws_summary.max_row]:
            c.font = Font(bold=True)
        created_sheets.append(ws_summary)

    if want_invoice:
        if want_summary:
            ws_invoice = wb.create_sheet("InvoiceWise")
        else:
            ws_invoice = wb.active
            ws_invoice.title = "InvoiceWise"
        _fill_dispatch_invoice_sheet(
            ws_invoice,
            client_name=client_name,
            po_no_val=po_no_val,
            invoice_rows=invoice_rows,
            title_font=title_font,
            header_fill=header_fill,
            header_font=header_font,
        )
        created_sheets.append(ws_invoice)

    for ws in created_sheets:
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe_client = re.sub(r"[^A-Za-z0-9._-]+", "_", client_name)
    safe_po = re.sub(r"[^A-Za-z0-9._-]+", "_", po_no_val)
    suffix = ""
    if want_summary and not want_invoice:
        suffix = "_summary"
    elif want_invoice and not want_summary:
        suffix = "_invoice_wise"
    filename = f"dispatch_{safe_client}_{safe_po}{suffix}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/ledger/export-xlsx")
def export_ledger_xlsx(
    payload: LedgerExportRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Excel export dependency missing. Install openpyxl.")
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access financial invoice data.")

    invoice_nos = [str(x or "").strip() for x in (payload.invoice_nos or []) if str(x or "").strip()]
    if not invoice_nos:
        raise HTTPException(status_code=400, detail="No invoice numbers provided for export.")

    invoices = db.query(Invoice).options(joinedload(Invoice.client), joinedload(Invoice.purchase_order)).filter(
        Invoice.invoice_no.in_(invoice_nos)
    ).all()
    if not invoices:
        raise HTTPException(status_code=404, detail="No matching invoices found for export.")

    invoice_map = {inv.invoice_no: inv for inv in invoices}
    ordered_invoices = [invoice_map[i] for i in invoice_nos if i in invoice_map]

    wb = Workbook()
    ws = wb.active
    ws.title = "FinancialLedger"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=12)

    ws["A1"] = f"Financial Ledger Export | Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:Q1")

    headers = [
        "Client", "Sub Entity", "PO No", "Invoice No", "Invoice Date", "Due Date", "Status",
        "LR No", "Basic", "GST", "Gross", "Advance", "TDS", "Retention", "Net Payable",
        "Payment Received", "Balance"
    ]
    ws.append(headers)
    for c in ws[2]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    today = datetime.date.today()
    totals = {
        "basic": 0.0, "gst": 0.0, "gross": 0.0, "advance": 0.0, "tds": 0.0,
        "retention": 0.0, "net": 0.0, "paid": 0.0, "balance": 0.0
    }
    for inv in ordered_invoices:
        inv_date = inv.inv_date.isoformat() if inv.inv_date else ""
        due_date = inv.due_date.isoformat() if inv.due_date else ""
        balance = float(inv.balance or 0.0)
        status = "CLEARED"
        if balance > 0.0001:
            if inv.due_date and inv.due_date < today:
                status = "OVERDUE"
            else:
                status = "PENDING"

        basic = float(inv.basic or 0.0)
        gst = float(inv.gst or 0.0)
        gross = float(inv.total or 0.0)
        advance = float(inv.advance_adj or 0.0)
        tds = float(inv.tds_ded or 0.0)
        retention = float(inv.retention_held or 0.0)
        net = float(inv.net_payable or 0.0)
        paid = float(inv.paid or 0.0)

        totals["basic"] += basic
        totals["gst"] += gst
        totals["gross"] += gross
        totals["advance"] += advance
        totals["tds"] += tds
        totals["retention"] += retention
        totals["net"] += net
        totals["paid"] += paid
        totals["balance"] += balance

        ws.append([
            inv.client.name if inv.client else "",
            inv.sub_entity or "",
            inv.purchase_order.po_no if inv.purchase_order else "UNASSIGNED",
            inv.invoice_no,
            inv_date,
            due_date,
            status,
            inv.lr_no or "",
            round(basic, 2),
            round(gst, 2),
            round(gross, 2),
            round(advance, 2),
            round(tds, 2),
            round(retention, 2),
            round(net, 2),
            round(paid, 2),
            round(balance, 2),
        ])

    ws.append([
        "TOTAL", "", "", "", "", "", "", "",
        round_inr_nearest(totals["basic"]),
        round_inr_nearest(totals["gst"]),
        round_inr_nearest(totals["gross"]),
        round_inr_nearest(totals["advance"]),
        round_inr_nearest(totals["tds"]),
        round_inr_nearest(totals["retention"]),
        round_inr_nearest(totals["net"]),
        round_inr_nearest(totals["paid"]),
        round_inr_nearest(totals["balance"]),
    ])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    widths = [24, 18, 16, 18, 14, 14, 12, 16, 12, 10, 12, 12, 10, 12, 12, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"financial_ledger_export_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/logistics/client-dispatch-workspace")
def get_logistics_client_dispatch_workspace(
    client: str = Query(..., description="Client name"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if current_user.role not in ["admin", "logistics"]:
        raise HTTPException(status_code=403, detail="Admin or Logistics access required")

    client_name = (client or "").strip()
    if not client_name:
        raise HTTPException(status_code=400, detail="client is required")

    db_client = db.query(Client).filter(Client.name == client_name).first()
    if not db_client:
        raise HTTPException(status_code=404, detail="Client not found")

    pos = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).filter(
        PurchaseOrder.client_id == db_client.id
    ).all()

    po_terms = []
    po_by_id = {}
    for po in pos:
        po_by_id[po.id] = po
        po_terms.append({
            "po_no": po.po_no,
            "project_name": po.project_name or "",
            "contact_person": po.contact_person or "",
            "is_completed": bool(po.is_completed),
            "is_hidden": bool(po.is_hidden),
            "completed_at": po.completed_at.isoformat() if po.completed_at else None,
            "baseline_items": [
                {
                    "description": item.description or "",
                    "ordered_qty": float(item.ordered_qty or 0.0),
                    "inspected_qty": float(item.inspected_qty or 0.0),
                    "uom": item.uom or "Nos",
                    "material_type": item.material_type or None,
                    "dispatch_alias": item.dispatch_alias or None,
                    "dispatch_rate": float(item.dispatch_rate or 0.0)
                } for item in (po.baseline_items or [])
            ]
        })

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.client_id == db_client.id,
        Invoice.is_note == False
    ).all()

    invoices_payload = []
    for inv in invoices:
        po_no = ""
        if inv.po_id and inv.po_id in po_by_id:
            po_no = po_by_id[inv.po_id].po_no
        invoices_payload.append({
            "id": inv.invoice_no,
            "invoice_no": inv.invoice_no,
            "po_no": po_no or "UNASSIGNED",
            "sub_entity": inv.sub_entity or "",
            "lr_no": inv.lr_no or "",
            "inv_date": inv.inv_date.isoformat() if inv.inv_date else None,
            "due_date": inv.due_date.isoformat() if inv.due_date else None,
            "total": float(inv.total or 0.0),
            "dispatch_items": [
                {
                    "id": d.id,
                    "description": d.description or "",
                    "qty": float(d.dispatched_qty or 0.0),
                    "inspected_qty": float(d.inspected_qty or 0.0),
                    "uom": d.uom or "Nos",
                    "rate_per_uom": float(d.rate_per_uom or 0.0)
                } for d in (inv.dispatch_items or [])
            ]
        })

    return {
        "client": {
            "id": db_client.id,
            "name": db_client.name,
            "active": bool(db_client.active if db_client.active is not None else True)
        },
        "po_terms": po_terms,
        "invoices": invoices_payload
    }



def normalize_dispatch_description(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]+", " ", (value or "").upper())).strip()


def prune_orphan_baseline_items_for_po(po_id: int, db: Session) -> int:
    po = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).filter(PurchaseOrder.id == po_id).first()
    if not po:
        return 0

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id == po_id,
        Invoice.is_note == False
    ).all()

    dispatch_qty_by_desc: dict[str, float] = defaultdict(float)
    for inv in invoices:
        for item in inv.dispatch_items:
            key = normalize_dispatch_description(item.description or "")
            if key:
                dispatch_qty_by_desc[key] += float(item.dispatched_qty or 0.0)

    deleted = 0
    for base in list(po.baseline_items or []):
        key = normalize_dispatch_description(base.description or "")
        ordered = float(base.ordered_qty or 0.0)
        inspected = float(base.inspected_qty or 0.0)
        dispatched = float(dispatch_qty_by_desc.get(key, 0.0))
        # Auto-generated/material-only rows should disappear once fully consumed.
        if ordered <= 0.0001 and inspected <= 0.0001 and dispatched <= 0.0001:
            db.delete(base)
            deleted += 1

    return deleted


def ensure_baseline_from_dispatch(po_id: Optional[int], dispatch_items: list[DispatchItemCreate], db: Session) -> int:
    if not po_id:
        return 0
    po = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).filter(PurchaseOrder.id == po_id).first()
    if not po:
        return 0

    existing_by_key: dict[tuple[str, str], PoBaselineItem] = {}
    for base in (po.baseline_items or []):
        key = (normalize_dispatch_description(base.description or ""), (base.uom or "").strip().upper())
        existing_by_key[key] = base

    inserted = 0
    for item in (dispatch_items or []):
        desc = (item.description or "").strip()
        if not desc:
            continue
        norm = normalize_dispatch_description(desc)
        uom = (item.uom or "Nos").strip()
        key = (norm, uom.upper())
        if key in existing_by_key:
            existing = existing_by_key[key]
            if float(item.rate_per_uom or 0.0) > 0:
                existing.dispatch_rate = float(item.rate_per_uom or 0.0)
            continue
        new_base = PoBaselineItem(
            po_id=po.id,
            description=desc,
            ordered_qty=0.0,
            inspected_qty=float(item.inspected_qty or 0.0),
            uom=uom,
            dispatch_rate=max(0.0, float(item.rate_per_uom or 0.0))
        )
        db.add(new_base)
        existing_by_key[key] = new_base
        inserted += 1
    return inserted


def best_dispatch_description_match(raw_desc: str, candidates: list[str]) -> Optional[str]:
    norm_raw = normalize_dispatch_description(raw_desc)
    if not norm_raw or not candidates:
        return None

    exact = next((c for c in candidates if normalize_dispatch_description(c) == norm_raw), None)
    if exact:
        return exact

    raw_tokens = set(norm_raw.split())
    best = None
    best_score = 0.0
    for candidate in candidates:
        norm_c = normalize_dispatch_description(candidate)
        if not norm_c:
            continue
        if norm_c in norm_raw or norm_raw in norm_c:
            score = min(len(norm_raw), len(norm_c)) / max(len(norm_raw), len(norm_c))
            if score > best_score:
                best_score = score
                best = candidate
            continue
        cand_tokens = norm_c.split()
        if not cand_tokens:
            continue
        common = sum(1 for t in cand_tokens if t in raw_tokens)
        score = common / max(1, min(len(cand_tokens), len(raw_tokens)))
        if score > best_score:
            best_score = score
            best = candidate
    return best if best_score >= 0.75 else None


def align_existing_dispatch_schema(db: Session) -> dict:
    po_list = db.query(PurchaseOrder).options(selectinload(PurchaseOrder.baseline_items)).all()
    updated_dispatch_rows = 0
    inserted_baseline_rows = 0
    scanned_dispatch_rows = 0

    for po in po_list:
        baseline_items = po.baseline_items or []
        baseline_by_norm = {}
        for b in baseline_items:
            key = normalize_dispatch_description(b.description or "")
            if key and key not in baseline_by_norm:
                baseline_by_norm[key] = b

        baseline_candidates = [b.description for b in baseline_items if b.description]
        invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
            Invoice.po_id == po.id,
            Invoice.is_note == False
        ).all()
        for inv in invoices:
            for item in inv.dispatch_items:
                scanned_dispatch_rows += 1
                source_desc = (item.description or "").strip()
                if not source_desc:
                    continue
                matched = best_dispatch_description_match(source_desc, baseline_candidates)
                if matched and source_desc != matched:
                    item.description = matched
                    updated_dispatch_rows += 1
                    source_desc = matched
                norm_desc = normalize_dispatch_description(source_desc)
                if norm_desc and norm_desc not in baseline_by_norm:
                    new_base = PoBaselineItem(
                        po_id=po.id,
                        description=source_desc,
                        ordered_qty=0.0,
                        inspected_qty=0.0,
                        uom=item.uom or "Nos",
                        dispatch_rate=max(0.0, float(getattr(item, "rate_per_uom", 0.0) or 0.0))
                    )
                    db.add(new_base)
                    baseline_by_norm[norm_desc] = new_base
                    baseline_candidates.append(source_desc)
                    inserted_baseline_rows += 1

    pruned_baseline_rows = 0
    for po in po_list:
        pruned_baseline_rows += prune_orphan_baseline_items_for_po(po.id, db)

    db.commit()
    return {
        "success": True,
        "po_count": len(po_list),
        "scanned_dispatch_rows": scanned_dispatch_rows,
        "dispatch_rows_aligned": updated_dispatch_rows,
        "baseline_rows_added": inserted_baseline_rows,
        "baseline_rows_pruned": pruned_baseline_rows
    }


@app.post("/api/dispatch/align-schema")
def align_dispatch_schema(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return align_existing_dispatch_schema(db)


def note_display_balance(inv: Invoice) -> float:
    """Signed Total Due for CN/DN rows (negative for credit notes)."""
    if is_debit_note(inv):
        return invoice_outstanding_balance(inv)
    stored = float(inv.balance or 0.0)
    total_num = float(inv.total or 0.0)
    ntype = str(inv.note_type or "").strip().upper()
    note_abs = abs(total_num)
    if ntype == "CN":
        if stored < -0.009:
            return stored
        if note_abs > 0.009:
            return total_num if total_num <= 0 else -note_abs
        return stored
    return stored


def is_debit_note(inv: Invoice) -> bool:
    return bool(inv.is_note) and str(inv.note_type or "").strip().upper() == "DN"


def invoice_eligible_for_payment_allocation(inv: Invoice) -> bool:
    """Regular invoices and debit notes can receive bank/unallocated allocations."""
    return not inv.is_note or is_debit_note(inv)


def invoice_outstanding_balance(inv: Invoice) -> float:
    """Amount still due on a row (matches frontend ledgerBalance for DN rows)."""
    if is_debit_note(inv):
        net = float(inv.net_payable or 0.0)
        paid = float(inv.paid or 0.0)
        if net > 0.009:
            return max(0.0, net - paid)
        stored = float(inv.balance or 0.0)
        if stored > 0.009:
            return stored
        note_amt = abs(float(inv.total or 0.0))
        if note_amt > 0.009:
            return note_amt
        return max(0.0, stored)
    return float(inv.balance or 0.0)


def sync_note_row_balance(inv: Invoice) -> None:
    """Set CN/DN row balances only; never changes linked invoice rows."""
    if not inv.is_note:
        return
    inv.advance_adj = 0.0
    inv.tds_ded = 0.0
    inv.retention_held = 0.0
    note_amt = abs(float(inv.total or 0.0))
    ntype = str(inv.note_type or "").strip().upper()
    if ntype == "CN":
        inv.net_payable = 0.0
        inv.paid = 0.0
        inv.balance = -note_amt
    elif ntype == "DN":
        inv.net_payable = note_amt
        inv.paid = 0.0
        inv.balance = note_amt
    else:
        inv.net_payable = 0.0
        inv.paid = 0.0
        inv.balance = float(inv.total or 0.0)


def recalculate_client_ledger(client_id: int, db: Session, preserve_manual_paid: bool = False):
    """
    The Master Math Engine: Calculates all invoices, deducts payments, 
    and locks the true balances directly into the SQL database.

    When ``preserve_manual_paid`` is True (used during invoice creation /
    legacy imports), an invoice's existing ``paid`` value is kept if it is
    larger than the sum of recorded allocations. This lets users seed
    historical invoices with an opening paid figure even before any
    payment record exists. Every payment lifecycle path (allocate /
    delete / redistribute / update / restore) keeps the default
    behaviour: ``paid`` is fully derived from current allocations.
    """
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client: return

    invoices = db.query(Invoice).filter(Invoice.client_id == client_id).all()
    # Create a fast lookup dictionary mapping invoice_no string to the SQL object
    inv_map = {inv.invoice_no: inv for inv in invoices}
    po_map: dict[int, PurchaseOrder] = {
        po.id: po for po in db.query(PurchaseOrder).filter(PurchaseOrder.client_id == client_id).all()
    }
    # Ensure PO advance wallet is always distributed to eligible invoices before
    # ledger math is recalculated (covers new invoices and term changes).
    _auto_apply_po_advance(client_id, db)
    db.flush()

    payments = db.query(PaymentHistory).filter(PaymentHistory.client_id == client_id).all()
    advance_applied_by_inv: dict[str, float] = defaultdict(float)
    alloc_paid_by_inv: dict[str, float] = defaultdict(float)
    payment_alloc_sum: dict[str, float] = defaultdict(float)
    for pay in payments:
        allocations = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == pay.id).all()
        for al in allocations:
            if al.alloc_type == 'invoice' and al.target_inv_id in inv_map:
                alloc_paid_by_inv[al.target_inv_id] += float(al.amount or 0.0)
            if al.alloc_type == 'po_advance_applied' and al.target_inv_id in inv_map:
                advance_applied_by_inv[al.target_inv_id] += float(al.amount or 0.0)
            if al.alloc_type in ('invoice', 'po_advance', 'po_advance_applied', 'note_allocation'):
                payment_alloc_sum[pay.id] += float(al.amount or 0.0)

    # 1. Reset all invoice balances to baseline.
    # Paid is allocation-driven (deleting a payment correctly re-opens the
    # invoice). The ``preserve_manual_paid`` flag is honoured below so legacy
    # imports that seed an opening ``paid`` figure aren't wiped out by the
    # post-create recalc when no payment record exists yet.
    for inv in invoices:
        if inv.is_note:
            ntype = str(inv.note_type or "").strip().upper()
            if ntype == "CN":
                sync_note_row_balance(inv)
                continue
            if ntype == "DN":
                note_amt = abs(float(inv.total or 0.0))
                inv.advance_adj = 0.0
                inv.tds_ded = 0.0
                inv.retention_held = 0.0
                inv.net_payable = note_amt
                # paid / balance are derived from receipt allocations in step 2.
                continue
            sync_note_row_balance(inv)
            continue

        manual_paid_pre = float(inv.paid or 0.0)
        po_obj = po_map.get(inv.po_id) if inv.po_id else None
        if po_obj is not None:
            ret_base_amount = invoice_amount_for_po_base(inv, po_obj.ret_base or "total")

            # Retention/advance follow PO terms; TDS is user-entered per invoice (never overwritten here).
            retention_pct = float(po_obj.ret_pct or 0.0)
            inv.retention_held = max(0.0, ret_base_amount * (retention_pct / 100.0))
            inv.tds_ded = max(0.0, float(inv.tds_ded or 0.0))

            adv_pct = float(po_obj.adv_pct or 0.0)
            max_adv_allowed = max(0.0, ret_base_amount * (adv_pct / 100.0)) if adv_pct > 0 else 0.0
            applied_adv = float(advance_applied_by_inv.get(inv.invoice_no, 0.0))
            inv.advance_adj = max(0.0, min(applied_adv, max_adv_allowed if adv_pct > 0 else 0.0))
        else:
            # Without a linked PO, retention / TDS / advance carry whatever the
            # invoice was created with (typical for UNASSIGNED placeholder POs).
            inv.advance_adj = float(inv.advance_adj or 0.0)
            inv.tds_ded = float(inv.tds_ded or 0.0)
            inv.retention_held = float(inv.retention_held or 0.0)

        # Balance due rule: gross less advance and TDS deduction.
        inv.net_payable = max(0.0, float(inv.total or 0.0) - float(inv.advance_adj or 0.0) - float(inv.tds_ded or 0.0))
        inv.paid = 0.0
        if preserve_manual_paid:
            inv._manual_paid_seed = manual_paid_pre  # type: ignore[attr-defined]
        inv.balance = float(inv.net_payable or 0.0)  # initialise; will be updated in step 2

    # 2. Layer allocation-derived payments onto the baseline.
    # NOTE: We intentionally do NOT cap derived_paid at net_payable. When the
    # user explicitly overpays an invoice (allow_overpayment=True on the
    # payment record), alloc_paid will exceed net_payable and the resulting
    # negative balance represents a credit on the account. Capping would
    # silently hide that credit and break pool-total accounting.
    total_excess = 0.0
    for inv in invoices:
        if inv.is_note and not is_debit_note(inv):
            continue
        alloc_paid = float(alloc_paid_by_inv.get(inv.invoice_no, 0.0))
        # CN rows are not paid via receipt allocation; DN rows are (like invoices).
        derived_paid = alloc_paid
        if preserve_manual_paid:
            manual_seed = float(getattr(inv, "_manual_paid_seed", 0.0) or 0.0)
            inv.paid = max(manual_seed, derived_paid)
        else:
            inv.paid = derived_paid
        inv.balance = float(inv.net_payable or 0.0) - float(inv.paid or 0.0)
        # balance can be negative for deliberately overpaid invoices (credit)

    # 3. Calculate Unallocated / Excess Funds
    for pay in payments:
        alloc_sum = float(payment_alloc_sum.get(pay.id, 0.0))
        if pay.type == 'RECEIPT':
            unallocated = pay.amount - alloc_sum
            if unallocated > 0:
                total_excess += unallocated
        elif pay.type == 'UNALLOCATED_APPLIED':
            # This log consumes previously accumulated unallocated funds.
            total_excess -= alloc_sum

    # Tiny residuals (< INR 5) are NOT auto-zeroed any more.
    # We keep the actual `paid` / `balance` numbers and the UI marks invoices
    # with balance < INR 5 as "CLEARED" (green) while still showing the real due,
    # so finance can see and reconcile the small leftovers if needed.

    client.excess_funds = round_inr_nearest(max(0.0, total_excess))

    # Persist PO wallet snapshot for dashboards / PO Advance Wallet.
    for po in db.query(PurchaseOrder).filter(PurchaseOrder.client_id == client_id).all():
        po.advance_pool = round_inr_nearest(max(0.0, _po_advance_pool_remaining_db(db, client_id, po.po_no)))

    db.commit()


def _po_advance_pool_remaining_db(db: Session, client_id: int, po_no: str) -> float:
    """Unapplied PO advance = sum(po_advance) - sum(po_advance_applied) for this PO."""
    rows = db.query(PaymentAllocation).join(
        PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
    ).filter(
        PaymentHistory.client_id == client_id,
        PaymentAllocation.target_po_no == po_no,
        PaymentAllocation.alloc_type.in_(["po_advance", "po_advance_applied"]),
    ).all()
    added = sum(float(r.amount or 0.0) for r in rows if r.alloc_type == "po_advance")
    used = sum(float(r.amount or 0.0) for r in rows if r.alloc_type == "po_advance_applied")
    return max(0.0, added - used)


def _strip_po_advance_applied_for_po(client_id: int, po_no: str, db: Session) -> int:
    """Remove every invoice advance application row for this PO (pool returns to unapplied)."""
    rows = db.query(PaymentAllocation).join(
        PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
    ).filter(
        PaymentHistory.client_id == client_id,
        PaymentAllocation.alloc_type == "po_advance_applied",
        PaymentAllocation.target_po_no == po_no,
    ).all()
    for r in rows:
        db.delete(r)
    return len(rows)


def _auto_apply_po_advance(client_id: int, db: Session, po_no: Optional[str] = None, invoice_no: Optional[str] = None):
    po_query = db.query(PurchaseOrder).filter(PurchaseOrder.client_id == client_id)
    if po_no:
        po_query = po_query.filter(PurchaseOrder.po_no == po_no)
    po_rows = po_query.all()
    if not po_rows:
        return

    for po in po_rows:
        adv_pct = float(po.adv_pct or 0.0)
        if adv_pct <= 0:
            _strip_po_advance_applied_for_po(client_id, po.po_no, db)
            continue

        # Build per-payment advance lots for this PO.
        added_by_payment: dict[str, float] = defaultdict(float)
        consumed_by_payment: dict[str, float] = defaultdict(float)
        po_allocs = db.query(PaymentAllocation).join(
            PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
        ).filter(
            PaymentHistory.client_id == client_id,
            PaymentAllocation.target_po_no == po.po_no,
            PaymentAllocation.alloc_type.in_(["po_advance", "po_advance_applied"]),
        ).all()
        for al in po_allocs:
            pid = str(al.payment_id or "").strip()
            if not pid:
                continue
            if al.alloc_type == "po_advance":
                added_by_payment[pid] += float(al.amount or 0.0)
            elif al.alloc_type == "po_advance_applied":
                consumed_by_payment[pid] += float(al.amount or 0.0)

        lots: list[list[object]] = []
        for pid, added_amt in added_by_payment.items():
            remaining_amt = float(added_amt) - float(consumed_by_payment.get(pid, 0.0))
            if remaining_amt > 0:
                lots.append([pid, remaining_amt])
        if not lots:
            continue
        lot_idx = 0

        invoices = db.query(Invoice).filter(
            Invoice.client_id == client_id,
            Invoice.po_id == po.id,
            Invoice.is_note == False
        ).all()
        invoices.sort(key=lambda inv: invoice_ledger_sort_key(inv.inv_date, inv.invoice_no))
        if invoice_no:
            invoices = [inv for inv in invoices if inv.invoice_no == invoice_no]
        if not invoices:
            continue

        applied_by_invoice: dict[str, float] = defaultdict(float)
        existing_applied = db.query(PaymentAllocation).join(
            PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
        ).filter(
            PaymentHistory.client_id == client_id,
            PaymentAllocation.alloc_type == "po_advance_applied",
            PaymentAllocation.target_po_no == po.po_no,
            PaymentAllocation.target_inv_id.isnot(None)
        ).all()
        for al in existing_applied:
            applied_by_invoice[str(al.target_inv_id)] += float(al.amount or 0.0)

        for inv in invoices:
            base_amt = float(inv.basic or 0.0) if (po.ret_base or "total") == "basic" else float(inv.total or 0.0)
            max_allowed = max(0.0, base_amt * (adv_pct / 100.0))
            current_applied = float(applied_by_invoice.get(inv.invoice_no, 0.0))
            shortfall = max(0.0, max_allowed - current_applied)
            if shortfall <= 0:
                continue

            while shortfall > 0 and lot_idx < len(lots):
                pid = str(lots[lot_idx][0])
                remaining_lot = float(lots[lot_idx][1])
                if remaining_lot <= 0:
                    lot_idx += 1
                    continue
                take = min(shortfall, remaining_lot)
                db.add(PaymentAllocation(
                    payment_id=pid,
                    alloc_type="po_advance_applied",
                    target_inv_id=inv.invoice_no,
                    target_po_no=po.po_no,
                    note_id=None,
                    amount=float(take),
                ))
                lots[lot_idx][1] = remaining_lot - take
                applied_by_invoice[inv.invoice_no] = applied_by_invoice.get(inv.invoice_no, 0.0) + take
                shortfall -= take


def round_inr_nearest(value: Optional[float]) -> float:
    amt = Decimal(str(float(value or 0.0)))
    return float(amt.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def invoice_amount_for_po_base(inv: Invoice, base_kind: Optional[str]) -> float:
    """Resolve PO term base ('basic' | 'total') to an invoice monetary amount."""
    kind = (base_kind or "total").strip().lower()
    if kind == "basic":
        return float(inv.basic or 0.0)
    return float(inv.total or 0.0)


def round_qty_total(value: Optional[float]) -> float:
    amt = Decimal(str(float(value or 0.0)))
    return float(amt.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def fiscal_year_label(inv_date: Optional[datetime.date], start_month: int, start_day: int) -> Optional[str]:
    if not inv_date:
        return None
    pivot = datetime.date(inv_date.year, start_month, start_day)
    start_year = inv_date.year if inv_date >= pivot else inv_date.year - 1
    return f"FY{start_year}-{str(start_year + 1)[-2:]}"


_INVOICE_FY_SEGMENT_RE = re.compile(r"^(\d{2})\s*-\s*(\d{2})$")


def invoice_no_ledger_sort_parts(invoice_no: Optional[str]) -> tuple[int, int, int, str]:
    """
    Parse invoice numbers like RS/25-26/123 for ledger ordering:
    financial-year segment (25-26) first, then trailing sequence (123).
    """
    raw = (invoice_no or "").strip()
    parts = [p.strip() for p in raw.split("/") if p.strip()]
    fy_start, fy_end, seq = 9999, 9999, 999_999_999
    if len(parts) >= 2:
        fy_match = _INVOICE_FY_SEGMENT_RE.match(parts[1])
        if fy_match:
            fy_start = int(fy_match.group(1))
            fy_end = int(fy_match.group(2))
    last_chunk = parts[-1] if parts else raw
    seq_match = re.search(r"(\d+)\s*$", last_chunk)
    if seq_match:
        seq = int(seq_match.group(1))
    return fy_start, fy_end, seq, raw.lower()


def invoice_ledger_sort_key(
    inv_date: Optional[datetime.date],
    invoice_no: Optional[str],
) -> tuple:
    """Primary: invoice date; secondary: FY segment then sequence from invoice no."""
    d = inv_date or datetime.date.max
    fy_start, fy_end, seq, raw = invoice_no_ledger_sort_parts(invoice_no)
    return (d, fy_start, fy_end, seq, raw)


def invoice_ledger_sort_key_from_strings(
    inv_date: Optional[str],
    invoice_no: Optional[str],
) -> tuple:
    d = inv_date or "9999-12-31"
    fy_start, fy_end, seq, raw = invoice_no_ledger_sort_parts(invoice_no)
    return (d, fy_start, fy_end, seq, raw)


def parse_fy_filters(raw: Optional[str]) -> set[str]:
    if not raw:
        return set()
    return {part.strip() for part in raw.split(",") if part.strip()}



@app.get("/api/invoices")
def get_invoices(fiscal_years: Optional[str] = Query(default=None, description="Comma separated FY labels"), include_completed_po: bool = True, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access financial invoice data.")
    invoices = db.query(Invoice).options(
        joinedload(Invoice.purchase_order),
        selectinload(Invoice.dispatch_items)
    ).all()
    settings = db.query(SystemSettings).first()
    fy_start_month = int((settings.fy_start_month if settings else 4) or 4)
    fy_start_day = int((settings.fy_start_day if settings else 1) or 1)
    fy_filters = parse_fy_filters(fiscal_years)
    note_invoice_nos = [inv.invoice_no for inv in invoices if inv.is_note]
    note_target_by_note_id: dict[str, str] = {}
    if note_invoice_nos:
        note_allocs = db.query(PaymentAllocation).filter(
            PaymentAllocation.alloc_type == "note_allocation",
            PaymentAllocation.note_id.in_(note_invoice_nos)
        ).all()
        for alloc in note_allocs:
            note_id = (alloc.note_id or "").strip()
            target_inv = (alloc.target_inv_id or "").strip()
            if note_id and target_inv and note_id not in note_target_by_note_id:
                note_target_by_note_id[note_id] = target_inv
        # Fallback for legacy/malformed rows where note_allocation linkage may
        # be missing but NOTE_APPLIED details still capture the target invoice.
        unresolved_note_nos = [n for n in note_invoice_nos if n not in note_target_by_note_id]
        if unresolved_note_nos:
            note_pays = db.query(PaymentHistory).filter(
                PaymentHistory.client_id.in_([inv.client_id for inv in invoices if inv.is_note]),
                PaymentHistory.type == "NOTE_APPLIED"
            ).all()
            note_apply_re = re.compile(r"\bNote\s+(.+?)\s+applied\s+to\s+(.+?)\s*$", re.IGNORECASE)
            for pay in note_pays:
                details = str(pay.details or "").strip()
                if not details:
                    continue
                m = note_apply_re.search(details)
                if not m:
                    continue
                note_id = m.group(1).strip()
                target_inv = m.group(2).strip()
                if note_id in unresolved_note_nos and target_inv and note_id not in note_target_by_note_id:
                    note_target_by_note_id[note_id] = target_inv
    result = []
    for inv in invoices:
        po_str = 'UNASSIGNED'
        po_completed = False
        if inv.purchase_order:
            po_str = inv.purchase_order.po_no
            po_completed = bool(inv.purchase_order.is_completed)
        elif inv.po_id:
            po_obj = db.query(PurchaseOrder).filter(PurchaseOrder.id == inv.po_id).first()
            if po_obj:
                po_str = po_obj.po_no
                po_completed = bool(po_obj.is_completed)
        # Fallback: for notes, derive PO from linked target invoice if this row
        # has no PO relation (guards legacy rows and partial data).
        if inv.is_note and po_str == 'UNASSIGNED':
            note_target = note_target_by_note_id.get(inv.invoice_no)
            if note_target:
                target_inv_obj = db.query(Invoice).options(joinedload(Invoice.purchase_order)).filter(
                    Invoice.invoice_no == note_target,
                    Invoice.client_id == inv.client_id
                ).first()
                if target_inv_obj:
                    if target_inv_obj.purchase_order:
                        po_str = target_inv_obj.purchase_order.po_no
                        po_completed = bool(target_inv_obj.purchase_order.is_completed)
                    elif target_inv_obj.po_id:
                        po_obj = db.query(PurchaseOrder).filter(PurchaseOrder.id == target_inv_obj.po_id).first()
                        if po_obj:
                            po_str = po_obj.po_no
                            po_completed = bool(po_obj.is_completed)
        if (not include_completed_po) and po_completed:
            continue
        inv_fy = fiscal_year_label(inv.inv_date, fy_start_month, fy_start_day)
        if fy_filters and inv_fy not in fy_filters:
            continue

        # --- FETCH ATTACHED DISPATCH ITEMS ---
        d_items = []
        for item in inv.dispatch_items:
            d_items.append({
                "id": item.id,
                "description": item.description,
                "qty": item.dispatched_qty,
                "inspected_qty": item.inspected_qty,
                "uom": item.uom,
                "rate_per_uom": float(item.rate_per_uom or 0.0)
            })

        result.append({
            "sql_id": inv.id,
            "client_id": inv.client_id,
            "poNo": po_str,
            "id": inv.invoice_no,
            "subEntity": inv.sub_entity,
            "lrNo": inv.lr_no,
            "invDate": inv.inv_date.isoformat() if inv.inv_date else None,
            "dueDate": inv.due_date.isoformat() if inv.due_date else None,
            "basic": inv.basic,
            "gst": inv.gst,
            "total": inv.total,
            "advance": inv.advance_adj,
            "tds": inv.tds_ded,
            "retention": inv.retention_held,
            "netPayable": inv.net_payable,
            "paid": inv.paid,
            "balance": note_display_balance(inv) if inv.is_note else inv.balance,
            "isNote": inv.is_note,
            "noteType": inv.note_type,
            "noteReason": inv.note_reason,
            "noteTargetInvoice": note_target_by_note_id.get(inv.invoice_no),
            "migratedV3": True,
            "fiscalYear": inv_fy,
            "dispatchItems": d_items # CRITICAL: Sends items to frontend memory
        })
    result.sort(key=lambda r: invoice_ledger_sort_key_from_strings(r.get("invDate"), r.get("id")))
    return result

@app.post("/api/invoices")
def create_invoice(inv: InvoiceCreate, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    
    existing = db.query(Invoice).filter(Invoice.invoice_no == inv.invoice_no).first()
    if existing:
        raise HTTPException(status_code=400, detail="Invoice number already exists.")

    po_id = None
    if inv.po_no and inv.po_no.strip() and inv.po_no != 'UNASSIGNED':
        po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == inv.po_no).first()
        if not po:
            po = PurchaseOrder(client_id=inv.client_id, po_no=inv.po_no)
            db.add(po)
            db.flush() 
        po_id = po.id

    inv_d = datetime.datetime.strptime(inv.inv_date, '%Y-%m-%d').date() if inv.inv_date else None
    due_d = datetime.datetime.strptime(inv.due_date, '%Y-%m-%d').date() if inv.due_date else None
    rounded_basic = round_inr_nearest(inv.basic)
    rounded_total = round_inr_nearest(inv.total)
    server_net_payable = max(
        0.0,
        float(rounded_total or 0.0) - float(inv.advance_adj or 0.0) - float(inv.tds_ded or 0.0),
    )
    server_balance = float(server_net_payable) - float(inv.paid or 0.0)

    new_inv = Invoice(
        client_id=inv.client_id, po_id=po_id, invoice_no=inv.invoice_no,
        sub_entity=inv.sub_entity, lr_no=inv.lr_no, inv_date=inv_d, due_date=due_d,
        basic=rounded_basic, gst=inv.gst, total=rounded_total, advance_adj=inv.advance_adj,
        tds_ded=inv.tds_ded, retention_held=inv.retention_held, net_payable=server_net_payable,
        paid=inv.paid, balance=server_balance, is_note=inv.is_note, note_type=inv.note_type, note_reason=inv.note_reason
    )
    db.add(new_inv)
    
    # CRITICAL FIX: Flush to generate new_inv.id BEFORE adding dispatch items
    db.flush() 

    for item in inv.dispatch_items:
        new_dispatch = InvoiceDispatchItem(
            invoice_id=new_inv.id,
            description=item.description,
            dispatched_qty=item.qty,
            inspected_qty=item.inspected_qty,
            uom=item.uom,
            rate_per_uom=max(0.0, float(item.rate_per_uom or 0.0))
        )
        db.add(new_dispatch)
    ensure_baseline_from_dispatch(po_id, inv.dispatch_items, db)

    record_audit(
        db,
        current_user,
        entity_type="invoice",
        entity_id=new_inv.invoice_no,
        action="create",
        summary=f"Created invoice {new_inv.invoice_no} ({rounded_total:.2f} total)",
        details={
            "client_id": new_inv.client_id,
            "po_no": inv.po_no,
            "basic": rounded_basic,
            "total": rounded_total,
            "dispatch_item_count": len(inv.dispatch_items or []),
        },
        request=request,
    )
    db.commit()
    auto_advance_applied = 0.0
    recalculate_client_ledger(new_inv.client_id, db, preserve_manual_paid=True)
    return {"success": True, "id": new_inv.id, "auto_advance_applied": auto_advance_applied}

@app.put("/api/invoices/{invoice_no:path}")
def update_invoice(invoice_no: str, inv: InvoiceUpdate, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    
    db_inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if not db_inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")
    before = {
        "basic": float(db_inv.basic or 0.0),
        "total": float(db_inv.total or 0.0),
        "advance_adj": float(db_inv.advance_adj or 0.0),
        "tds_ded": float(db_inv.tds_ded or 0.0),
        "retention_held": float(db_inv.retention_held or 0.0),
        "paid": float(db_inv.paid or 0.0),
        "balance": float(db_inv.balance or 0.0),
        "po_id": db_inv.po_id,
    }
        
    po_id = None
    if inv.po_no and inv.po_no.strip() and inv.po_no != 'UNASSIGNED':
        po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == inv.po_no).first()
        if not po:
            po = PurchaseOrder(client_id=inv.client_id, po_no=inv.po_no)
            db.add(po)
            db.flush() 
        po_id = po.id

    db_inv.po_id = po_id
    db_inv.sub_entity = inv.sub_entity
    db_inv.lr_no = inv.lr_no
    db_inv.inv_date = datetime.datetime.strptime(inv.inv_date, '%Y-%m-%d').date() if inv.inv_date else None
    db_inv.due_date = datetime.datetime.strptime(inv.due_date, '%Y-%m-%d').date() if inv.due_date else None
    db_inv.basic = round_inr_nearest(inv.basic)
    db_inv.gst = inv.gst
    db_inv.total = round_inr_nearest(inv.total)
    db_inv.advance_adj = inv.advance_adj
    db_inv.tds_ded = inv.tds_ded
    db_inv.retention_held = inv.retention_held
    db_inv.net_payable = max(
        0.0,
        float(db_inv.total or 0.0) - float(db_inv.advance_adj or 0.0) - float(db_inv.tds_ded or 0.0),
    )
    db_inv.paid = inv.paid
    db_inv.balance = float(db_inv.net_payable or 0.0) - float(db_inv.paid or 0.0)
    db_inv.is_note = inv.is_note
    db_inv.note_type = inv.note_type
    db_inv.note_reason = inv.note_reason
    
    # CRITICAL FIX: Clear old items and write new ones safely using db_inv.id
    db.query(InvoiceDispatchItem).filter(InvoiceDispatchItem.invoice_id == db_inv.id).delete()
    for item in inv.dispatch_items:
        new_dispatch = InvoiceDispatchItem(
            invoice_id=db_inv.id,
            description=item.description,
            dispatched_qty=item.qty,
            inspected_qty=item.inspected_qty,
            uom=item.uom,
            rate_per_uom=max(0.0, float(item.rate_per_uom or 0.0))
        )
        db.add(new_dispatch)
    ensure_baseline_from_dispatch(po_id, inv.dispatch_items, db)

    after = {
        "basic": float(db_inv.basic or 0.0),
        "total": float(db_inv.total or 0.0),
        "advance_adj": float(db_inv.advance_adj or 0.0),
        "tds_ded": float(db_inv.tds_ded or 0.0),
        "retention_held": float(db_inv.retention_held or 0.0),
        "paid": float(db_inv.paid or 0.0),
        "balance": float(db_inv.balance or 0.0),
        "po_id": db_inv.po_id,
    }
    record_audit(
        db,
        current_user,
        entity_type="invoice",
        entity_id=db_inv.invoice_no,
        action="update",
        summary=f"Updated invoice {db_inv.invoice_no}",
        details={"before": before, "after": after, "po_no": inv.po_no},
        request=request,
    )
    db.commit()
    recalculate_client_ledger(db_inv.client_id, db)
    return {"success": True}

@app.delete("/api/invoices/{invoice_no:path}")
def delete_invoice(invoice_no: str, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    db_inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if db_inv:
        client_id = db_inv.client_id
        snapshot = {
            "client_id": db_inv.client_id,
            "po_id": db_inv.po_id,
            "basic": float(db_inv.basic or 0.0),
            "total": float(db_inv.total or 0.0),
            "paid": float(db_inv.paid or 0.0),
            "balance": float(db_inv.balance or 0.0),
            "is_note": bool(db_inv.is_note),
        }
        # Track how much receipt money gets reopened as unallocated because of this delete.
        reopened_unallocated_by_payment: dict[str, float] = defaultdict(float)

        # Remove every allocation linked to this invoice (all alloc types), so if the
        # same invoice number is added again it starts clean without retained payment map.
        allocs = db.query(PaymentAllocation).filter(
            or_(
                PaymentAllocation.target_inv_id == db_inv.invoice_no,
                PaymentAllocation.note_id == db_inv.invoice_no
            )
        ).all()
        pay_map = {}
        if allocs:
            pids = list({a.payment_id for a in allocs if a.payment_id})
            if pids:
                pay_rows = db.query(PaymentHistory).filter(PaymentHistory.id.in_(pids)).all()
                pay_map = {p.id: p for p in pay_rows}
        linked_payment_ids = {a.payment_id for a in allocs if a.payment_id}
        for al in allocs:
            pay_obj = pay_map.get(al.payment_id)
            if pay_obj and pay_obj.type == "RECEIPT" and al.alloc_type == "invoice":
                reopened_unallocated_by_payment[al.payment_id] += float(al.amount or 0.0)
            db.delete(al)
        # Cleanup any auto-generated payment logs that no longer have allocations.
        for pid in linked_payment_ids:
            remaining = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == pid).count()
            if remaining == 0:
                ph = db.query(PaymentHistory).filter(
                    PaymentHistory.id == pid,
                    PaymentHistory.type.in_(["ADVANCE_APPLIED", "NOTE_APPLIED", "UNALLOCATED_APPLIED"])
                ).first()
                if ph:
                    db.delete(ph)
        db.flush()
        po_no = db_inv.purchase_order.po_no if db_inv.purchase_order else "UNASSIGNED"
        for pay_id, reopened_amt in reopened_unallocated_by_payment.items():
            reopened_amt = round_inr_nearest(reopened_amt)
            if reopened_amt <= 0:
                continue
            db.add(UnallocatedPaymentRegister(
                client_id=client_id,
                source_payment_id=pay_id,
                created_on=datetime.date.today(),
                amount=float(reopened_amt),
                balance=float(reopened_amt),
                status="open",
                note=_build_unallocated_register_note(
                    "invoice_deleted",
                    f"Invoice {invoice_no} deleted; allocation reopened.",
                    invoice_no=invoice_no,
                    po_no=po_no,
                ),
            ))
        deleted_po_no = db_inv.purchase_order.po_no if db_inv.purchase_order else None
        db.delete(db_inv)
        db.flush()
        if deleted_po_no:
            _auto_apply_po_advance(client_id, db, deleted_po_no)
        record_audit(
            db,
            current_user,
            entity_type="invoice",
            entity_id=invoice_no,
            action="delete",
            summary=f"Deleted invoice {invoice_no}",
            details=snapshot,
            request=request,
        )
        db.commit()
        _auto_apply_po_advance(client_id, db)
        recalculate_client_ledger(client_id, db)
    return {"success": True}


@app.delete("/api/dispatch-items/{dispatch_item_id}")
def delete_dispatch_item(dispatch_item_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics"]:
        raise HTTPException(status_code=403, detail="Admin or Logistics access required")
    item = db.query(InvoiceDispatchItem).filter(InvoiceDispatchItem.id == dispatch_item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Dispatch entry not found.")
    invoice = db.query(Invoice).filter(Invoice.id == item.invoice_id).first()
    po_id = invoice.po_id if invoice else None
    db.delete(item)
    db.flush()
    pruned = 0
    if po_id:
        pruned = prune_orphan_baseline_items_for_po(po_id, db)
    db.commit()
    return {"success": True, "baseline_rows_pruned": pruned}


@app.put("/api/dispatch-items/{dispatch_item_id}")
def update_dispatch_item(dispatch_item_id: int, payload: DispatchItemUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")

    item = db.query(InvoiceDispatchItem).filter(InvoiceDispatchItem.id == dispatch_item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Dispatch entry not found.")

    invoice = db.query(Invoice).filter(Invoice.id == item.invoice_id).first()
    po_id = invoice.po_id if invoice else None

    desc = (payload.description or "").strip()
    if not desc:
        raise HTTPException(status_code=400, detail="Description is required.")

    item.description = desc
    item.dispatched_qty = max(0.0, float(payload.qty or 0.0))
    item.inspected_qty = max(0.0, float(payload.inspected_qty or 0.0))
    item.uom = (payload.uom or "Nos").strip() or "Nos"

    merged = False
    merged_into_id = None
    if payload.merge_on_match:
        norm_desc = normalize_dispatch_description(item.description)
        norm_uom = (item.uom or "").strip().upper()
        siblings = db.query(InvoiceDispatchItem).filter(
            InvoiceDispatchItem.invoice_id == item.invoice_id,
            InvoiceDispatchItem.id != item.id
        ).all()
        for sibling in siblings:
            if normalize_dispatch_description(sibling.description or "") == norm_desc and (sibling.uom or "").strip().upper() == norm_uom:
                sibling.dispatched_qty = float(sibling.dispatched_qty or 0.0) + float(item.dispatched_qty or 0.0)
                sibling.inspected_qty = float(sibling.inspected_qty or 0.0) + float(item.inspected_qty or 0.0)
                merged = True
                merged_into_id = sibling.id
                db.delete(item)
                break

    if po_id:
        ensure_baseline_from_dispatch(
            po_id,
            [DispatchItemCreate(description=desc, qty=max(0.0, float(payload.qty or 0.0)), inspected_qty=max(0.0, float(payload.inspected_qty or 0.0)), uom=(payload.uom or "Nos"))],
            db
        )
        prune_orphan_baseline_items_for_po(po_id, db)

    db.commit()
    return {"success": True, "merged": merged, "merged_into_id": merged_into_id}


@app.post("/api/dispatch-items/upsert-cell")
def upsert_dispatch_cell(payload: DispatchCellUpsert, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")

    invoice_no = (payload.invoice_no or "").strip()
    desc = (payload.description or "").strip()
    if not invoice_no:
        raise HTTPException(status_code=400, detail="invoice_no is required.")
    if not desc:
        raise HTTPException(status_code=400, detail="description is required.")

    inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    qty = max(0.0, float(payload.qty or 0.0))
    uom = (payload.uom or "Nos").strip() or "Nos"
    norm_desc = normalize_dispatch_description(desc)
    norm_uom = uom.upper()

    matches = db.query(InvoiceDispatchItem).filter(InvoiceDispatchItem.invoice_id == inv.id).all()
    matched = [m for m in matches if normalize_dispatch_description(m.description or "") == norm_desc and (m.uom or "").strip().upper() == norm_uom]

    if matched:
        keeper = matched[0]
        keeper.description = desc
        keeper.uom = uom
        keeper.dispatched_qty = qty
        if float(payload.rate_per_uom or 0.0) > 0:
            keeper.rate_per_uom = float(payload.rate_per_uom or 0.0)
        # Keep inspected qty bounded to dispatched qty.
        keeper.inspected_qty = min(float(keeper.inspected_qty or 0.0), qty)
        for extra in matched[1:]:
            db.delete(extra)
        dispatch_item_id = keeper.id
        created = False
    else:
        new_item = InvoiceDispatchItem(
            invoice_id=inv.id,
            description=desc,
            dispatched_qty=qty,
            inspected_qty=0.0,
            uom=uom,
            rate_per_uom=max(0.0, float(payload.rate_per_uom or 0.0))
        )
        db.add(new_item)
        db.flush()
        dispatch_item_id = new_item.id
        created = True

    if inv.po_id:
        ensure_baseline_from_dispatch(
            inv.po_id,
            [DispatchItemCreate(description=desc, qty=qty, inspected_qty=0.0, uom=uom, rate_per_uom=float(payload.rate_per_uom or 0.0))],
            db
        )
        prune_orphan_baseline_items_for_po(inv.po_id, db)

    db.commit()
    return {"success": True, "created": created, "dispatch_item_id": dispatch_item_id}


@app.patch("/api/invoices/{invoice_no:path}/inline")
def update_invoice_inline_fields(invoice_no: str, payload: InvoiceInlineUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")

    if payload.inv_date is not None:
        inv.inv_date = datetime.datetime.strptime(payload.inv_date, "%Y-%m-%d").date() if str(payload.inv_date).strip() else None
    if payload.lr_no is not None:
        inv.lr_no = (payload.lr_no or "").strip()
    if payload.total is not None:
        inv.total = max(0.0, round_inr_nearest(payload.total))
        # Keep payable consistency when total is edited inline.
        inv.net_payable = max(
            0.0,
            float(inv.total or 0.0) - float(inv.advance_adj or 0.0) - float(inv.tds_ded or 0.0),
        )
        inv.balance = float(inv.net_payable or 0.0) - float(inv.paid or 0.0)

    db.commit()
    recalculate_client_ledger(inv.client_id, db)
    return {"success": True}


@app.delete("/api/dispatch-row/{invoice_no:path}")
def delete_dispatch_row(invoice_no: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found.")
    po_id = inv.po_id
    deleted = db.query(InvoiceDispatchItem).filter(InvoiceDispatchItem.invoice_id == inv.id).delete()
    if po_id:
        prune_orphan_baseline_items_for_po(po_id, db)
    db.commit()
    return {"success": True, "deleted_dispatch_lines": int(deleted or 0)}


@app.post("/api/dispatch-columns/delete")
def delete_dispatch_column(payload: DispatchColumnDeleteRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    po_no = (payload.po_no or "").strip()
    desc = (payload.description or "").strip()
    if not po_no or not desc:
        raise HTTPException(status_code=400, detail="po_no and description are required.")
    po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first()
    if not po:
        raise HTTPException(status_code=404, detail="PO not found.")

    norm_desc = normalize_dispatch_description(desc)
    norm_uom = (payload.uom or "").strip().upper() if payload.uom else ""

    baseline_deleted = 0
    for base in list(po.baseline_items or []):
        if normalize_dispatch_description(base.description or "") != norm_desc:
            continue
        if norm_uom and (base.uom or "").strip().upper() != norm_uom:
            continue
        db.delete(base)
        baseline_deleted += 1

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id == po.id,
        Invoice.is_note == False
    ).all()
    dispatch_deleted = 0
    for inv in invoices:
        for item in list(inv.dispatch_items or []):
            if normalize_dispatch_description(item.description or "") != norm_desc:
                continue
            if norm_uom and (item.uom or "").strip().upper() != norm_uom:
                continue
            db.delete(item)
            dispatch_deleted += 1

    prune_orphan_baseline_items_for_po(po.id, db)
    db.commit()
    return {"success": True, "baseline_deleted": baseline_deleted, "dispatch_deleted": dispatch_deleted}


@app.post("/api/dispatch-columns/rename")
def rename_dispatch_column(payload: DispatchColumnRenameRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Rename a dispatch material column on a PO and migrate every dispatch_item that
    used the old description to the new one. If the new description matches an
    existing column for the same UOM, quantities are merged so multiple items
    can share the same material description."""
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    po_no = (payload.po_no or "").strip()
    old_desc_raw = (payload.old_description or "").strip()
    new_desc_raw = (payload.new_description or "").strip()
    if not po_no or not old_desc_raw or not new_desc_raw:
        raise HTTPException(status_code=400, detail="po_no, old_description and new_description are required.")

    po = db.query(PurchaseOrder).filter(PurchaseOrder.po_no == po_no).first()
    if not po:
        raise HTTPException(status_code=404, detail="PO not found.")

    old_norm = normalize_dispatch_description(old_desc_raw)
    new_norm = normalize_dispatch_description(new_desc_raw)
    old_uom = (payload.old_uom or "Nos").strip() or "Nos"
    new_uom = (payload.new_uom or old_uom or "Nos").strip() or "Nos"
    old_uom_key = old_uom.upper()
    new_uom_key = new_uom.upper()

    baseline_renamed = 0
    baseline_merged = 0
    keeper_baseline = None
    target_baseline = None
    other_baselines: list = []
    for base in list(po.baseline_items or []):
        norm = normalize_dispatch_description(base.description or "")
        u = (base.uom or "Nos").strip().upper()
        if norm == new_norm and u == new_uom_key:
            keeper_baseline = base
        elif norm == old_norm and u == old_uom_key:
            other_baselines.append(base)

    if keeper_baseline is None and other_baselines:
        target_baseline = other_baselines.pop(0)
        target_baseline.description = new_desc_raw
        target_baseline.uom = new_uom
        if hasattr(target_baseline, "material_type") and not target_baseline.material_type:
            target_baseline.material_type = target_baseline.material_type
        keeper_baseline = target_baseline
        baseline_renamed += 1

    if keeper_baseline is not None:
        for base in other_baselines:
            keeper_baseline.ordered_qty = float(keeper_baseline.ordered_qty or 0.0) + float(base.ordered_qty or 0.0)
            keeper_baseline.inspected_qty = float(keeper_baseline.inspected_qty or 0.0) + float(base.inspected_qty or 0.0)
            if not keeper_baseline.material_type and getattr(base, "material_type", None):
                keeper_baseline.material_type = base.material_type
            if not getattr(keeper_baseline, "dispatch_alias", None) and getattr(base, "dispatch_alias", None):
                keeper_baseline.dispatch_alias = base.dispatch_alias
            db.delete(base)
            baseline_merged += 1

    invoices = db.query(Invoice).options(selectinload(Invoice.dispatch_items)).filter(
        Invoice.po_id == po.id,
        Invoice.is_note == False
    ).all()
    dispatch_renamed = 0
    dispatch_merged = 0
    for inv in invoices:
        keeper_item = None
        renames: list = []
        for item in list(inv.dispatch_items or []):
            norm = normalize_dispatch_description(item.description or "")
            u = (item.uom or "Nos").strip().upper()
            if norm == new_norm and u == new_uom_key:
                keeper_item = item
            elif norm == old_norm and u == old_uom_key:
                renames.append(item)
        if not renames:
            continue
        if keeper_item is None:
            keeper_item = renames.pop(0)
            keeper_item.description = new_desc_raw
            keeper_item.uom = new_uom
            dispatch_renamed += 1
        for extra in renames:
            keeper_item.dispatched_qty = float(keeper_item.dispatched_qty or 0.0) + float(extra.dispatched_qty or 0.0)
            keeper_item.inspected_qty = float(keeper_item.inspected_qty or 0.0) + float(extra.inspected_qty or 0.0)
            db.delete(extra)
            dispatch_merged += 1

    prune_orphan_baseline_items_for_po(po.id, db)
    db.commit()
    return {
        "success": True,
        "baseline_renamed": baseline_renamed,
        "baseline_merged": baseline_merged,
        "dispatch_renamed": dispatch_renamed,
        "dispatch_merged": dispatch_merged,
    }


def _normalize_source_po_nos(values: list[str] | None) -> set[str]:
    return {str(v or "").strip() for v in (values or []) if str(v or "").strip()}


def _source_po_for_unallocated_register(db: Session, reg: UnallocatedPaymentRegister) -> str:
    _source, inv_no, po_no, _rest = _parse_unallocated_register_note(reg.note)
    if po_no:
        return po_no

    source_pos: set[str] = set()
    if inv_no:
        inv = db.query(Invoice).filter(
            Invoice.client_id == reg.client_id,
            Invoice.invoice_no == inv_no,
        ).first()
        if inv and inv.po_id:
            po = db.query(PurchaseOrder).filter(PurchaseOrder.id == inv.po_id).first()
            if po and po.po_no:
                source_pos.add(po.po_no)

    if reg.source_payment_id:
        allocations = db.query(PaymentAllocation).filter(
            PaymentAllocation.payment_id == reg.source_payment_id
        ).all()
        inv_ids = [a.target_inv_id for a in allocations if a.target_inv_id]
        if inv_ids:
            rows = (
                db.query(Invoice, PurchaseOrder)
                .outerjoin(PurchaseOrder, Invoice.po_id == PurchaseOrder.id)
                .filter(Invoice.client_id == reg.client_id, Invoice.invoice_no.in_(inv_ids))
                .all()
            )
            for _inv, po in rows:
                source_pos.add(po.po_no if po and po.po_no else "UNASSIGNED")
        for alloc in allocations:
            if alloc.target_po_no:
                source_pos.add(alloc.target_po_no)

    if len(source_pos) == 1:
        return next(iter(source_pos))
    if len(source_pos) > 1:
        return "MULTI"
    return "UNASSIGNED"


def _open_unallocated_registers(
    db: Session,
    client_id: int,
    source_po_nos: set[str] | None = None,
) -> list[UnallocatedPaymentRegister]:
    regs = (
        db.query(UnallocatedPaymentRegister)
        .filter(
            UnallocatedPaymentRegister.client_id == client_id,
            UnallocatedPaymentRegister.status == "open",
            UnallocatedPaymentRegister.balance > 0,
        )
        .order_by(UnallocatedPaymentRegister.id.asc())
        .all()
    )
    if not source_po_nos:
        return regs
    return [reg for reg in regs if _source_po_for_unallocated_register(db, reg) in source_po_nos]


def _available_unallocated_register_balance(
    db: Session,
    client_id: int,
    source_po_nos: set[str] | None = None,
) -> float:
    return round_inr_nearest(
        sum(float(reg.balance or 0.0) for reg in _open_unallocated_registers(db, client_id, source_po_nos))
    )


@app.post("/api/payments/allocate")
def allocate_payment(payment: PaymentAllocateRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if payment.advance_only and not payment.apply_adv:
        raise HTTPException(status_code=400, detail="advance_only requires apply_adv=true.")
    if payment.fund_source not in ("receipt", "unallocated"):
        raise HTTPException(status_code=400, detail="fund_source must be 'receipt' or 'unallocated'.")
    if payment.fund_source == "receipt" and not payment.advance_only and payment.amount <= 0:
        raise HTTPException(status_code=400, detail="Payment amount must be greater than zero.")
    if payment.clear_po_pool and not payment.po_no:
        raise HTTPException(status_code=400, detail="po_no is required when clear_po_pool=true.")
    if db.query(PaymentHistory).filter(PaymentHistory.id == payment.id).first():
        raise HTTPException(status_code=400, detail="Payment id already exists.")

    date_obj = datetime.datetime.strptime(payment.date, '%Y-%m-%d').date()
    invoices = db.query(Invoice).filter(Invoice.client_id == payment.client_id).all()
    invoices = [inv for inv in invoices if invoice_eligible_for_payment_allocation(inv)]
    inv_map = {inv.invoice_no: inv for inv in invoices}
    po_by_id = {po.id: po for po in db.query(PurchaseOrder).filter(PurchaseOrder.client_id == payment.client_id).all()}

    selected: list[tuple[Invoice, float]] = []
    if payment.mode == "targeted" and payment.targets:
        for t in payment.targets:
            inv = inv_map.get(t.inv_id)
            if not inv:
                continue
            req_amt = max(0.0, float(t.amount or 0.0))
            outstanding = invoice_outstanding_balance(inv)
            if payment.advance_only:
                if outstanding > 0.009 or req_amt > 0.009:
                    selected.append((inv, -1.0))
            elif req_amt > 0.009:
                selected.append((inv, req_amt))
            elif outstanding > 0.009:
                selected.append((inv, -1.0))
    else:
        selected = [(inv, -1.0) for inv in sorted(
            [i for i in invoices if invoice_outstanding_balance(i) > 0.009],
            key=lambda x: invoice_ledger_sort_key(x.inv_date, x.invoice_no),
        )]

    client = db.query(Client).filter(Client.id == payment.client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    source_po_nos = _normalize_source_po_nos(payment.source_po_nos)
    available_unallocated = (
        _available_unallocated_register_balance(db, payment.client_id, source_po_nos)
        if source_po_nos and payment.fund_source == "unallocated"
        else float(client.excess_funds or 0.0)
    )
    if payment.fund_source == "unallocated":
        if available_unallocated <= 0:
            raise HTTPException(status_code=400, detail="No unallocated funds available.")
        if payment.amount > 0:
            remaining = min(float(payment.amount), available_unallocated)
        else:
            remaining = available_unallocated
    else:
        remaining = float(payment.amount)

    advance_applied_total = 0.0
    po_advance_allocs: list[tuple[str, float, str]] = []
    allocs_for_db = []
    log_details = []

    po_pool: dict[str, float] = defaultdict(float)
    historical_allocs = db.query(PaymentAllocation).join(PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id).filter(
        PaymentHistory.client_id == payment.client_id
    ).all()
    for al in historical_allocs:
        if not al.target_po_no:
            continue
        if al.alloc_type == 'po_advance':
            po_pool[al.target_po_no] += float(al.amount or 0.0)
        elif al.alloc_type == 'po_advance_applied':
            po_pool[al.target_po_no] -= float(al.amount or 0.0)
    for po_no in list(po_pool.keys()):
        if po_pool[po_no] < 0:
            po_pool[po_no] = 0.0

    if payment.clear_po_pool:
        po_key = payment.po_no.strip()
        amount_to_clear = float(po_pool.get(po_key, 0.0))
        if amount_to_clear <= 0:
            raise HTTPException(status_code=400, detail="No remaining PO advance found to clear.")
        po_advance_allocs.append(("WRITTEN_OFF", amount_to_clear, po_key))
    elif payment.fund_source == "receipt" and payment.move_to_po and payment.move_to_po.strip():
        move_po = payment.move_to_po.strip()
        amount_to_move = remaining
        if amount_to_move <= 0:
            raise HTTPException(status_code=400, detail="Amount to move must be greater than zero.")
        allocs_for_db.append(("__PO__", amount_to_move))
        remaining = 0.0
    elif payment.fund_source == "unallocated" and payment.move_to_po and payment.move_to_po.strip():
        move_po = payment.move_to_po.strip()
        amount_to_move = remaining
        if amount_to_move <= 0:
            raise HTTPException(status_code=400, detail="Amount to move must be greater than zero.")
        allocs_for_db.append(("__PO__", amount_to_move))
        remaining = 0.0
    else:
        if payment.po_no and payment.po_no.strip():
            scoped_po = payment.po_no.strip()
            selected = [
                (inv, req) for inv, req in selected
                if (po_by_id.get(inv.po_id).po_no if inv.po_id and po_by_id.get(inv.po_id) else None) == scoped_po
            ]
        for inv, requested in selected:
            if payment.apply_adv and not payment.only_gst and not inv.is_note:
                po_obj = po_by_id.get(inv.po_id) if inv.po_id else None
                po_no = po_obj.po_no if po_obj else None
                adv_pct = float(po_obj.adv_pct or 0.0) if po_obj else 0.0
                base_key = (po_obj.ret_base or "total") if po_obj else "total"
                available_pool = float(po_pool.get(po_no, 0.0)) if po_no else 0.0
                if po_no and adv_pct > 0 and available_pool > 0:
                    base_amt = float(inv.basic or 0.0) if base_key == "basic" else float(inv.total or 0.0)
                    max_allowed = base_amt * (adv_pct / 100.0)
                    current_advance = float(inv.advance_adj or 0.0)
                    shortfall = max(0.0, max_allowed - current_advance)
                    inv_balance_for_adv = float(inv.balance or 0.0)
                    to_apply_adv = min(shortfall, available_pool, inv_balance_for_adv)
                    if to_apply_adv > 0:
                        inv.advance_adj = current_advance + to_apply_adv
                        po_pool[po_no] = available_pool - to_apply_adv
                        po_advance_allocs.append((inv.invoice_no, to_apply_adv, po_no))
                        advance_applied_total += to_apply_adv

            if payment.advance_only:
                continue
            if remaining <= 0:
                break
            inv_balance = invoice_outstanding_balance(inv)

            # Overpayment: only honoured when mode=targeted and a specific
            # positive amount was requested for this invoice. Cascade mode
            # never overpays (it stops when balance reaches 0).
            allow_overpay_this = (
                payment.allow_overpayment
                and payment.mode == "targeted"
                and requested > 0
            )

            if not allow_overpay_this and inv_balance <= 0.009:
                continue

            desired = remaining if requested < 0 else requested

            if allow_overpay_this:
                # No allocatable cap — the user explicitly chose to overpay.
                allocatable = float(desired)
            elif is_debit_note(inv):
                # DN rows have no GST/retention buckets — apply to full outstanding due.
                allocatable = inv_balance
            elif payment.only_gst:
                allocatable = min(float(inv.gst or 0.0), inv_balance)
            else:
                target_bal = 0.0
                if payment.hold_ret:
                    target_bal += float(inv.retention_held or 0.0)
                if payment.hold_gst:
                    target_bal += float(inv.gst or 0.0)
                allocatable = max(0.0, inv_balance - target_bal)

            if allocatable <= 0 and not allow_overpay_this:
                continue

            amount_to_apply = min(float(desired), allocatable, remaining)
            if amount_to_apply <= 0:
                continue

            allocs_for_db.append((inv.invoice_no, amount_to_apply))
            log_details.append(f"{inv.invoice_no} ({amount_to_apply:.2f})")
            remaining -= amount_to_apply

    if remaining > 0 and payment.fund_source == "receipt" and payment.excess_action == "allocate_pending":
        selected_ids = {inv.invoice_no for inv, _ in selected}
        pending_invoices = sorted(
            [
                i for i in invoices
                if invoice_outstanding_balance(i) > 0.009 and i.invoice_no not in selected_ids
            ],
            key=lambda x: invoice_ledger_sort_key(x.inv_date, x.invoice_no),
        )
        for inv in pending_invoices:
            if remaining <= 0:
                break
            inv_balance = invoice_outstanding_balance(inv)
            if inv_balance <= 0.009:
                continue
            amount_to_apply = min(inv_balance, remaining)
            allocs_for_db.append((inv.invoice_no, amount_to_apply))
            log_details.append(f"{inv.invoice_no} ({amount_to_apply:.2f})")
            remaining -= amount_to_apply

    if payment.clear_po_pool:
        payment_type = 'ADVANCE_APPLIED'
        payment_amount = po_advance_allocs[0][1] if po_advance_allocs else 0.0
    elif payment.advance_only:
        payment_type = 'ADVANCE_APPLIED'
        payment_amount = advance_applied_total
    elif payment.fund_source == "unallocated":
        payment_type = 'UNALLOCATED_APPLIED'
        payment_amount = (available_unallocated - remaining) if payment.amount <= 0 else min(float(payment.amount), available_unallocated) - remaining
    else:
        payment_type = 'RECEIPT'
        payment_amount = float(payment.amount)
    details = ", ".join(log_details) if log_details else ("PO advance mapping" if payment.advance_only else "Unallocated receipt")

    new_pay = PaymentHistory(
        id=payment.id,
        client_id=payment.client_id,
        date=date_obj,
        type=payment_type,
        amount=payment_amount,
        details=details,
        note=payment.note
    )
    db.add(new_pay)
    db.flush()

    if payment.fund_source == "receipt" and remaining > 0:
        db.add(UnallocatedPaymentRegister(
            client_id=payment.client_id,
            source_payment_id=new_pay.id,
            created_on=date_obj,
            amount=float(remaining),
            balance=float(remaining),
            status="open",
            note=_build_unallocated_register_note("direct_payment", payment.note)
        ))

    for inv_no, amt in allocs_for_db:
        if inv_no == "__PO__":
            moved_amt = float(amt or 0.0)
            if moved_amt > 0:
                db.add(UnallocatedAdvanceRegister(
                    client_id=payment.client_id,
                    source_payment_id=new_pay.id,
                    po_no=payment.move_to_po.strip() if payment.move_to_po else None,
                    created_on=date_obj,
                    amount=moved_amt,
                    balance=moved_amt,
                    status="open",
                    note=payment.note
                ))
            db.add(PaymentAllocation(
                payment_id=new_pay.id,
                alloc_type='po_advance',
                target_inv_id=None,
                target_po_no=payment.move_to_po.strip() if payment.move_to_po else None,
                note_id=None,
                amount=amt
            ))
        else:
            db.add(PaymentAllocation(
                payment_id=new_pay.id,
                alloc_type='invoice',
                target_inv_id=inv_no,
                target_po_no=None,
                note_id=None,
                amount=amt
            ))

    for inv_no, amt, po_no in po_advance_allocs:
        db.add(PaymentAllocation(
            payment_id=new_pay.id,
            alloc_type='po_advance_applied',
            target_inv_id=inv_no,
            target_po_no=po_no,
            note_id=None,
            amount=amt
        ))

    # Whenever PO advance is parked, consume it against all eligible PO invoices automatically.
    if payment.move_to_po and payment.move_to_po.strip():
        db.flush()
        _auto_apply_po_advance(payment.client_id, db, payment.move_to_po.strip())

    if payment.fund_source == "unallocated" and payment_amount > 0:
        _drain_unallocated_register(db, payment.client_id, payment_amount, source_po_nos=source_po_nos)

    db.commit()
    recalculate_client_ledger(payment.client_id, db)
    bank_allocated = 0.0 if (payment.advance_only or payment.fund_source == "unallocated") else (float(payment.amount) - remaining)
    return {
        "success": True,
        "allocated": bank_allocated,
        "po_advance_applied": advance_applied_total,
        "remaining": remaining,
        "allocation_count": len(allocs_for_db),
        "advance_allocation_count": len(po_advance_allocs),
        "unallocated_consumed": payment_amount if payment.fund_source == "unallocated" else 0.0
    }

@app.get("/api/payments")
def get_payments(fiscal_years: Optional[str] = Query(default=None, description="Comma separated FY labels"), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access payment data.")
    payments = db.query(PaymentHistory).options(selectinload(PaymentHistory.allocations)).all()
    settings = db.query(SystemSettings).first()
    fy_start_month = int((settings.fy_start_month if settings else 4) or 4)
    fy_start_day = int((settings.fy_start_day if settings else 1) or 1)
    fy_filters = parse_fy_filters(fiscal_years)

    # Build a lookup so allocations can be enriched with the note type (CN/DN)
    # without N+1 queries; the SPA's Payment Log cell uses this to colour and
    # sign the trace line.
    note_type_by_no: dict[str, Optional[str]] = {
        i.invoice_no: i.note_type
        for i in db.query(Invoice.invoice_no, Invoice.note_type).filter(Invoice.is_note.is_(True)).all()
    }

    result = []

    for p in payments:
        pay_fy = fiscal_year_label(p.date, fy_start_month, fy_start_day)
        if fy_filters and pay_fy not in fy_filters:
            continue
        alloc_list = []
        for a in p.allocations:
            alloc_list.append({
                "type": a.alloc_type,
                "invId": a.target_inv_id,
                "po": a.target_po_no,
                "noteId": a.note_id,
                "noteType": note_type_by_no.get(a.note_id) if a.note_id else None,
                "amount": a.amount
            })
            
        result.append({
            "id": p.id,
            "client_id": p.client_id,
            "date": p.date.isoformat() if p.date else None,
            "type": p.type,
            "amount": p.amount,
            "details": p.details,
            "note": p.note,
            "fiscalYear": pay_fy,
            "allocations": alloc_list
        })
        
    return result


def _build_unallocated_register_note(source_kind: str, raw_note: Optional[str], invoice_no: str = "", po_no: str = "") -> str:
    parts = [f"SRC={source_kind}"]
    if invoice_no:
        parts.append(f"INV={invoice_no}")
    if po_no:
        parts.append(f"PO={po_no}")
    payload = "|".join(parts)
    tail = (raw_note or "").strip()
    return f"[{payload}] {tail}".strip()


def _drain_unallocated_register(
    db: Session,
    client_id: int,
    amount: float,
    *,
    source_po_nos: set[str] | None = None,
) -> None:
    """Consume open register balances FIFO when drawing from the unallocated pool."""
    to_deduct = round_inr_nearest(max(0.0, float(amount or 0.0)))
    if to_deduct <= 0:
        return
    open_regs = _open_unallocated_registers(db, client_id, source_po_nos)
    for reg in open_regs:
        if to_deduct <= 0:
            break
        reg_bal = float(reg.balance or 0.0)
        take = min(reg_bal, to_deduct)
        reg.balance = round_inr_nearest(reg_bal - take)
        if reg.balance < 0.01:
            reg.balance = 0.0
            reg.status = "used"
        to_deduct = round_inr_nearest(to_deduct - take)
    if to_deduct > 0.009:
        scope = f" for PO filter {', '.join(sorted(source_po_nos))}" if source_po_nos else ""
        raise HTTPException(status_code=400, detail=f"Insufficient unallocated register balance{scope}.")


def _restore_unallocated_register_consumption(
    db: Session,
    client_id: int,
    amount: float,
    *,
    note: str = "",
    source_payment_id: Optional[str] = None,
) -> float:
    """Credit register rows LIFO after reversing a UNALLOCATED_APPLIED payment."""
    remaining = round_inr_nearest(max(0.0, float(amount or 0.0)))
    if remaining <= 0:
        return 0.0

    regs = (
        db.query(UnallocatedPaymentRegister)
        .filter(
            UnallocatedPaymentRegister.client_id == client_id,
            UnallocatedPaymentRegister.status.in_(["used", "open"]),
        )
        .order_by(UnallocatedPaymentRegister.id.desc())
        .all()
    )
    for reg in regs:
        if remaining <= 0:
            break
        headroom = round_inr_nearest(float(reg.amount or 0.0) - float(reg.balance or 0.0))
        if headroom < 0.01:
            continue
        add = min(headroom, remaining)
        reg.balance = round_inr_nearest(float(reg.balance or 0.0) + add)
        if reg.balance >= float(reg.amount or 0.0) - 0.01:
            reg.balance = float(reg.amount or 0.0)
        reg.status = "open"
        remaining = round_inr_nearest(remaining - add)

    if remaining > 0.01:
        db.add(
            UnallocatedPaymentRegister(
                client_id=client_id,
                source_payment_id=source_payment_id,
                created_on=datetime.date.today(),
                amount=float(remaining),
                balance=float(remaining),
                status="open",
                note=_build_unallocated_register_note(
                    "pool_restored",
                    note or "Reversed unallocated allocation",
                ),
            )
        )
        remaining = 0.0
    return float(amount or 0.0)


def _parse_unallocated_register_note(note: Optional[str]) -> tuple[str, Optional[str], Optional[str], str]:
    txt = str(note or "").strip()
    if not (txt.startswith("[") and "]" in txt):
        return "direct_payment", None, None, txt
    meta = txt[1:txt.index("]")]
    rest = txt[txt.index("]") + 1 :].strip()
    source = "direct_payment"
    inv = None
    po = None
    for token in meta.split("|"):
        k, _, v = token.partition("=")
        key = (k or "").strip().upper()
        val = (v or "").strip()
        if key == "SRC" and val:
            source = "invoice_deleted" if val.lower() == "invoice_deleted" else "direct_payment"
        elif key == "INV" and val:
            inv = val
        elif key == "PO" and val:
            po = val
    return source, inv, po, rest


@app.get("/api/registers/unallocated-payments")
def get_unallocated_payment_register(client_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access payment data.")
    rows = db.query(UnallocatedPaymentRegister).filter(
        UnallocatedPaymentRegister.client_id == client_id
    ).order_by(UnallocatedPaymentRegister.id.desc()).all()
    result = []
    for r in rows:
        src, inv_no, po_no, clean_note = _parse_unallocated_register_note(r.note)
        result.append({
            "id": r.id,
            "created_on": r.created_on.isoformat() if r.created_on else None,
            "amount": float(r.amount or 0.0),
            "balance": float(r.balance or 0.0),
            "status": r.status,
            "source_payment_id": r.source_payment_id,
            "source_kind": src,
            "source_invoice_no": inv_no,
            "source_po_no": po_no,
            "note": clean_note
        })
    return result


@app.post("/api/registers/unallocated-payments/{entry_id}/reverse")
def reverse_unallocated_payment_entry(entry_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    row = db.query(UnallocatedPaymentRegister).filter(UnallocatedPaymentRegister.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Register entry not found.")
    row.status = "reversed"
    row.balance = 0.0
    db.commit()
    recalculate_client_ledger(row.client_id, db)
    return {"success": True}


@app.get("/api/registers/unallocated-advances")
def get_unallocated_advance_register(client_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access financial advance data.")
    rows = db.query(UnallocatedAdvanceRegister).filter(
        UnallocatedAdvanceRegister.client_id == client_id
    ).order_by(UnallocatedAdvanceRegister.id.desc()).all()
    return [{
        "id": r.id,
        "created_on": r.created_on.isoformat() if r.created_on else None,
        "amount": float(r.amount or 0.0),
        "balance": float(r.balance or 0.0),
        "status": r.status,
        "po_no": r.po_no,
        "source_payment_id": r.source_payment_id,
        "note": r.note
    } for r in rows]


@app.post("/api/registers/unallocated-advances/{entry_id}/reverse")
def reverse_unallocated_advance_entry(entry_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    row = db.query(UnallocatedAdvanceRegister).filter(UnallocatedAdvanceRegister.id == entry_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Register entry not found.")
    row.status = "reversed"
    row.balance = 0.0
    db.commit()
    recalculate_client_ledger(row.client_id, db)
    return {"success": True}

@app.delete("/api/payments/{payment_id}")
def delete_payment(payment_id: str, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db_pay = db.query(PaymentHistory).filter(PaymentHistory.id == payment_id).first()
    if db_pay:
        client_id = db_pay.client_id
        if db_pay.type == "UNALLOCATED_APPLIED":
            _restore_unallocated_register_consumption(
                db,
                client_id,
                float(db_pay.amount or 0.0),
                note=db_pay.note or "",
                source_payment_id=payment_id,
            )
        linked_allocs = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == payment_id).all()
        linked_unalloc_pay = db.query(UnallocatedPaymentRegister).filter(
            UnallocatedPaymentRegister.source_payment_id == payment_id
        ).all()
        linked_unalloc_adv = db.query(UnallocatedAdvanceRegister).filter(
            UnallocatedAdvanceRegister.source_payment_id == payment_id
        ).all()
        snapshot = {
            "client_id": db_pay.client_id,
            "type": db_pay.type,
            "amount": float(db_pay.amount or 0.0),
            "details": db_pay.details,
            "allocation_count": len(linked_allocs),
        }
        po_nos_to_reset = {
            str(al.target_po_no).strip()
            for al in linked_allocs
            if al.alloc_type == "po_advance" and al.target_po_no and str(al.target_po_no).strip()
        }
        for row in linked_unalloc_pay:
            db.delete(row)
        for row in linked_unalloc_adv:
            db.delete(row)
        for al in linked_allocs:
            db.delete(al)
        db.delete(db_pay)
        record_audit(
            db,
            current_user,
            entity_type="payment",
            entity_id=payment_id,
            action="delete",
            summary=f"Deleted payment {payment_id} ({snapshot['amount']:.2f})",
            details=snapshot,
            request=request,
        )
        db.commit()
        # Any PO advance parked on this payment: strip all advance applications on that PO
        # so invoices deallocate, then re-apply from surviving advance lots only.
        for pono in po_nos_to_reset:
            _strip_po_advance_applied_for_po(client_id, pono, db)
        db.commit()
        _auto_apply_po_advance(client_id, db)
        recalculate_client_ledger(client_id, db)
    return {"success": True}

@app.post("/api/payments/{payment_id}/redistribute")
def redistribute_payment(payment_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    db_pay = db.query(PaymentHistory).options(selectinload(PaymentHistory.allocations)).filter(PaymentHistory.id == payment_id).first()
    if not db_pay:
        raise HTTPException(status_code=404, detail="Payment not found")

    response_payload = {
        "id": db_pay.id,
        "type": db_pay.type,
        "amount": float(db_pay.amount or 0.0),
        "note": db_pay.note or ""
    }
    client_id = db_pay.client_id
    if db_pay.type == "UNALLOCATED_APPLIED":
        _restore_unallocated_register_consumption(
            db,
            client_id,
            float(db_pay.amount or 0.0),
            note=db_pay.note or "",
            source_payment_id=payment_id,
        )
    linked_allocs = db.query(PaymentAllocation).filter(PaymentAllocation.payment_id == payment_id).all()
    po_nos_to_reset = {
        str(al.target_po_no).strip()
        for al in linked_allocs
        if al.alloc_type == "po_advance" and al.target_po_no and str(al.target_po_no).strip()
    }
    for row in db.query(UnallocatedPaymentRegister).filter(UnallocatedPaymentRegister.source_payment_id == payment_id).all():
        db.delete(row)
    for row in db.query(UnallocatedAdvanceRegister).filter(UnallocatedAdvanceRegister.source_payment_id == payment_id).all():
        db.delete(row)
    for al in linked_allocs:
        db.delete(al)
    db.delete(db_pay)
    db.commit()
    for pono in po_nos_to_reset:
        _strip_po_advance_applied_for_po(client_id, pono, db)
    db.commit()
    _auto_apply_po_advance(client_id, db)
    recalculate_client_ledger(client_id, db)
    return {"success": True, "payment": response_payload}

@app.put("/api/payments/{payment_id}")
def update_payment(payment_id: str, pay_update: PaymentUpdate, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    db_pay = db.query(PaymentHistory).filter(PaymentHistory.id == payment_id).first()
    if not db_pay:
        raise HTTPException(status_code=404, detail="Payment not found")
    before_amount = float(db_pay.amount or 0.0)
    before_note = db_pay.note
    db_pay.amount = pay_update.amount
    db_pay.note = pay_update.note
    record_audit(
        db,
        current_user,
        entity_type="payment",
        entity_id=payment_id,
        action="update",
        summary=f"Updated payment {payment_id}",
        details={
            "before": {"amount": before_amount, "note": before_note},
            "after": {"amount": float(db_pay.amount or 0.0), "note": db_pay.note},
        },
        request=request,
    )
    db.commit()
    recalculate_client_ledger(db_pay.client_id, db)
    return {"success": True}


@app.post("/api/clients/{client_id}/ledger/recalculate")
def recalculate_client_ledger_manual(
    client_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Force full PO-advance distribution + ledger math for this client (admin safety valve)."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    record_audit(
        db,
        current_user,
        entity_type="client",
        entity_id=str(client_id),
        action="ledger_recalculate",
        summary=f"Manual ledger recalculation for client {client_id}",
        details={},
        request=request,
    )
    db.commit()
    _auto_apply_po_advance(client_id, db)
    recalculate_client_ledger(client_id, db)
    return {"success": True}


@app.get("/api/clients/{client_id}/po-advance-pools")
def get_po_advance_pools(client_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role == "logistics":
        raise HTTPException(status_code=403, detail="Logistics role cannot access financial advance data.")
    if not db.query(Client).filter(Client.id == client_id).first():
        raise HTTPException(status_code=404, detail="Client not found")

    pos = db.query(PurchaseOrder).filter(PurchaseOrder.client_id == client_id).order_by(PurchaseOrder.po_no.asc()).all()
    pools_out: list[dict] = []

    for po in pos:
        po_allocs = db.query(PaymentAllocation).join(
            PaymentHistory, PaymentAllocation.payment_id == PaymentHistory.id
        ).filter(
            PaymentHistory.client_id == client_id,
            PaymentAllocation.target_po_no == po.po_no,
            PaymentAllocation.alloc_type.in_(["po_advance", "po_advance_applied"]),
        ).all()

        lot_map: dict[str, dict[str, float]] = defaultdict(lambda: {"parked": 0.0, "applied": 0.0})
        for al in po_allocs:
            pid = str(al.payment_id or "")
            if not pid:
                continue
            if al.alloc_type == "po_advance":
                lot_map[pid]["parked"] += float(al.amount or 0.0)
            elif al.alloc_type == "po_advance_applied":
                lot_map[pid]["applied"] += float(al.amount or 0.0)

        lots: list[dict] = []
        for pid, vals in sorted(lot_map.items()):
            ph = db.query(PaymentHistory).filter(PaymentHistory.id == pid).first()
            lots.append({
                "payment_id": pid,
                "payment_date": ph.date.isoformat() if ph and ph.date else None,
                "payment_type": ph.type if ph else None,
                "parked": round(float(vals["parked"]), 2),
                "applied_from_lot": round(float(vals["applied"]), 2),
                "lot_remaining": round(max(0.0, float(vals["parked"]) - float(vals["applied"])), 2),
            })

        inv_rows: list[dict] = []
        invoices = db.query(Invoice).filter(
            Invoice.client_id == client_id,
            Invoice.po_id == po.id,
            Invoice.is_note == False,
        ).all()
        invoices.sort(key=lambda inv: invoice_ledger_sort_key(inv.inv_date, inv.invoice_no))
        adv_pct = float(po.adv_pct or 0.0)
        base_key = (po.ret_base or "total")
        applied_by_inv = defaultdict(float)
        for al in po_allocs:
            if al.alloc_type == "po_advance_applied" and al.target_inv_id:
                applied_by_inv[str(al.target_inv_id)] += float(al.amount or 0.0)
        for inv in invoices:
            base_amt = float(inv.basic or 0.0) if base_key == "basic" else float(inv.total or 0.0)
            max_allowed = round(max(0.0, base_amt * (adv_pct / 100.0)), 2) if adv_pct > 0 else 0.0
            allocated = round(float(applied_by_inv.get(inv.invoice_no, 0.0)), 2)
            inv_rows.append({
                "invoice_no": inv.invoice_no,
                "inv_date": inv.inv_date.isoformat() if inv.inv_date else None,
                "total": float(inv.total or 0.0),
                "advance_on_invoice": float(inv.advance_adj or 0.0),
                "allocated_from_pool": allocated,
                "max_per_po_terms": max_allowed,
                "shortfall": round(max(0.0, max_allowed - allocated), 2),
            })

        pool_remaining = round(_po_advance_pool_remaining_db(db, client_id, po.po_no), 2)
        pools_out.append({
            "po_no": po.po_no,
            "adv_pct": adv_pct,
            "ret_base": base_key,
            "pool_remaining": pool_remaining,
            "advance_pool_column": round(float(po.advance_pool or 0.0), 2),
            "lots": lots,
            "invoices": inv_rows,
        })

    return {"client_id": client_id, "pools": pools_out}


@app.post("/api/clients/{client_id}/po-advance/manual-apply")
def manual_apply_po_advance_pool(
    client_id: int,
    payload: PoAdvanceManualApplyRequest,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Apply PO advance pool to invoices per current terms (all PO invoices, or one invoice if specified)."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if not db.query(Client).filter(Client.id == client_id).first():
        raise HTTPException(status_code=404, detail="Client not found")
    po = db.query(PurchaseOrder).filter(
        PurchaseOrder.client_id == client_id,
        PurchaseOrder.po_no == payload.po_no.strip(),
    ).first()
    if not po:
        raise HTTPException(status_code=404, detail="PO not found for this client")

    record_audit(
        db,
        current_user,
        entity_type="purchase_order",
        entity_id=payload.po_no,
        action="po_advance_manual_apply",
        summary=f"Manual PO advance apply {payload.po_no}",
        details={"invoice_no": payload.invoice_no},
        request=request,
    )
    db.commit()

    inv_filter = (payload.invoice_no or "").strip() or None
    _auto_apply_po_advance(client_id, db, payload.po_no.strip(), inv_filter)
    db.commit()
    recalculate_client_ledger(client_id, db)
    pool_after = round(_po_advance_pool_remaining_db(db, client_id, payload.po_no.strip()), 2)
    return {"success": True, "pool_remaining": pool_after}


@app.post("/api/invoices/{invoice_no:path}/transfer")
def transfer_invoice(invoice_no: str, req: TransferRequest, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    db_inv = db.query(Invoice).filter(Invoice.invoice_no == invoice_no).first()
    if not db_inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    old_client_id = db_inv.client_id
    record_audit(
        db,
        current_user,
        entity_type="invoice",
        entity_id=invoice_no,
        action=f"transfer:{req.action}",
        summary=f"{req.action.capitalize()} invoice {invoice_no} from client {old_client_id} to {req.new_client_id}",
        details={"action": req.action, "from_client_id": old_client_id, "to_client_id": req.new_client_id},
        request=request,
    )

    if req.action == "copy":
        new_inv = Invoice(
            invoice_no=f"{db_inv.invoice_no}-COPY", client_id=req.new_client_id, po_id=None,
            sub_entity=db_inv.sub_entity, lr_no=db_inv.lr_no, inv_date=db_inv.inv_date, due_date=db_inv.due_date,
            basic=db_inv.basic, gst=db_inv.gst, total=db_inv.total, advance_adj=0, tds_ded=db_inv.tds_ded,
            retention_held=db_inv.retention_held, net_payable=db_inv.net_payable, paid=0, balance=db_inv.net_payable,
            is_note=db_inv.is_note, note_type=db_inv.note_type, note_reason=db_inv.note_reason
        )
        db.add(new_inv)
        db.flush()
        for item in db_inv.dispatch_items:
            new_item = InvoiceDispatchItem(
                invoice_id=new_inv.id, description=item.description, dispatched_qty=item.dispatched_qty,
                inspected_qty=item.inspected_qty, uom=item.uom, rate_per_uom=float(item.rate_per_uom or 0.0)
            )
            db.add(new_item)
        db.commit()
        recalculate_client_ledger(req.new_client_id, db)
    elif req.action == "move":
        db_inv.client_id = req.new_client_id
        db_inv.po_id = None
        db_inv.advance_adj = 0
        db_inv.paid = 0
        db.commit()
        recalculate_client_ledger(old_client_id, db)
        recalculate_client_ledger(req.new_client_id, db)
        
    return {"success": True}

@app.post("/api/notes/issue")
def issue_note(req: NoteIssueRequest, request: Request, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if req.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than zero.")
    if req.note_type not in ("CN", "DN"):
        raise HTTPException(status_code=400, detail="Invalid note type.")
    if db.query(Invoice).filter(Invoice.invoice_no == req.note_no).first():
        raise HTTPException(status_code=400, detail="Document Number already exists.")

    note_date = datetime.datetime.strptime(req.date, '%Y-%m-%d').date() if req.date else datetime.date.today()
    total_amt = -req.amount if req.note_type == "CN" else req.amount
    target_invoice_id = (req.target_invoice_id or "").strip() or None
    target_invoice = None
    note_po_id = None
    note_target_po_no = None
    if target_invoice_id:
        target_invoice = db.query(Invoice).filter(
            Invoice.invoice_no == target_invoice_id,
            Invoice.client_id == req.client_id
        ).first()
        if not target_invoice:
            raise HTTPException(status_code=404, detail="Target invoice not found for the selected client.")
        note_po_id = target_invoice.po_id
        if target_invoice.purchase_order:
            note_target_po_no = target_invoice.purchase_order.po_no
        elif note_po_id:
            po_obj = db.query(PurchaseOrder).filter(PurchaseOrder.id == note_po_id).first()
            if po_obj:
                note_target_po_no = po_obj.po_no

    new_note = Invoice(
        client_id=req.client_id,
        po_id=note_po_id,
        invoice_no=req.note_no,
        sub_entity='-',
        lr_no='-',
        inv_date=note_date,
        due_date=note_date,
        basic=0.0,
        gst=0.0,
        total=total_amt,
        advance_adj=0.0,
        tds_ded=0.0,
        retention_held=0.0,
        net_payable=0.0,
        paid=0.0,
        balance=0.0,
        is_note=True,
        note_type=req.note_type,
        note_reason=req.reason
    )
    db.add(new_note)

    if target_invoice_id:
        pay_id = str(int(datetime.datetime.now(datetime.timezone.utc).timestamp() * 1000))
        db_pay = PaymentHistory(
            id=pay_id,
            client_id=req.client_id,
            date=note_date,
            type='NOTE_APPLIED',
            amount=float(req.amount),
            details=f"{'Credit' if req.note_type == 'CN' else 'Debit'} Note {req.note_no} applied to {target_invoice_id}",
            note=req.reason
        )
        db.add(db_pay)
        db.flush()
        db.add(PaymentAllocation(
            payment_id=db_pay.id,
            alloc_type='note_allocation',
            target_inv_id=target_invoice_id,
            target_po_no=note_target_po_no,
            note_id=req.note_no,
            amount=float(req.amount)
        ))

    record_audit(
        db,
        current_user,
        entity_type="note",
        entity_id=req.note_no,
        action=f"issue:{req.note_type}",
        summary=f"Issued {req.note_type} {req.note_no} ({req.amount:.2f}) against invoice {target_invoice_id or 'N/A'}",
        details={
            "client_id": req.client_id,
            "note_type": req.note_type,
            "amount": float(req.amount),
            "target_invoice_id": target_invoice_id,
            "target_po_no": note_target_po_no,
            "reason": req.reason,
        },
        request=request,
    )
    db.commit()
    db.refresh(new_note)
    # Do not run full client ledger recalc here — it can change the target invoice's
    # advance/TDS/retention from PO rules. Notes are separate ledger rows only.
    sync_note_row_balance(new_note)
    db.commit()
    return {"success": True}


# --- Domain routers (Workstream 2A carve-out) ---
# Imports happen here, AFTER every helper / model symbol the routers reference
# is defined. Adding a new router is a single line; removing one is a single
# line. Keep this block close to the bottom of the file.
from routers.audit import router as _audit_router
from routers.reconciliation import router as _reconciliation_router
from routers.health import router as _health_router
app.include_router(_audit_router)
app.include_router(_reconciliation_router)
app.include_router(_health_router)


# ... (All your schemas and API endpoints must be ABOVE this point) ...

# --- AI PDF Extraction Route ---

INVOICE_EXTRACTION_PROMPT = """
You are an expert invoice data extraction assistant. Extract details from this invoice
and return them as STRICT JSON (no markdown, no commentary, no backticks).

CRITICAL INSTRUCTION FOR ITEMS:
  - For 'desc' you MUST capture the ENTIRE multi-line description for each line.
  - Every line item MUST include a numeric 'rate' representing the per-unit price (rate per UOM).
  - If the invoice prints Amount (line total) and Qty but not Rate, compute rate = amount / qty
    and ALSO populate 'line_amount' with the printed line total when visible.
  - Use plain numbers (no currency symbols, no thousands separators).

CRITICAL INSTRUCTION FOR CLIENT:
  - 'clientName' must be the BUYER / CONSIGNEE / "Bill To" company name (NOT the seller).
  - Return the full registered company name exactly as printed.
  - If not present, return an empty string.

Return JSON exactly matching this structure:
{
  "clientName": "Buyer / Bill-To company name exactly as printed",
  "invNo": "Invoice Number",
  "poNo": "PO Number",
  "lrNo": "LR Number",
  "date": "YYYY-MM-DD",
  "basic": 1234.50,
  "items": [
    {
      "desc": "ENTIRE paragraph of the goods description exactly as written",
      "qty": 10.5,
      "uom": "MT",
      "rate": 123.45,
      "line_amount": 1296.23
    }
  ]
}
"""


def _build_invoice_generation_config():
    """Return a generation_config dict that asks Gemini for valid JSON."""

    config: dict[str, Any] = {"response_mime_type": "application/json"}
    return config


def _extract_text_from_genai_response(response: Any) -> str:
    """Plain text / JSON string from Gemini (legacy or google-genai).

    ``GenerateContentResponse.text`` is normally set, but some responses only
    populate ``candidates[0].content.parts[*].text`` (e.g. mixed modalities or
    edge SDK paths). An empty `.text` with non-empty parts caused the SPA to see
    ``raw_data === ''`` and show 'No valid invoice payload could be extracted.'
    """

    if response is None:
        return ""

    blob = getattr(response, "text", None)
    if blob is not None and str(blob).strip():
        return str(blob)

    parts = getattr(response, "parts", None)
    if parts:
        chunks = [str(getattr(p, "text", None) or "") for p in parts if getattr(p, "text", None)]
        merged = "".join(chunks).strip()
        if merged:
            return merged

    try:
        cands = getattr(response, "candidates", None) or []
        if cands and getattr(cands[0], "content", None):
            plist = getattr(cands[0].content, "parts", None) or []
            chunks = [
                str(getattr(p, "text", None) or "") for p in plist if getattr(p, "text", None)
            ]
            merged = "".join(chunks).strip()
            if merged:
                return merged
    except Exception:
        pass

    return ""


async def _gemini_extract_one_invoice(
    model,
    file_name: str,
    content: bytes,
    semaphore: asyncio.Semaphore,
) -> dict:
    """Extract a single invoice PDF with bounded concurrency and JSON normalization."""

    file_part = {"mime_type": "application/pdf", "data": content}
    generation_config = _build_invoice_generation_config()

    async with semaphore:
        try:
            response = await asyncio.to_thread(
                model.generate_content,
                [INVOICE_EXTRACTION_PROMPT, file_part],
                generation_config=generation_config,
            )
        except TypeError:
            # Older SDKs may not accept generation_config kwarg; fall back to plain call.
            try:
                response = await asyncio.to_thread(
                    model.generate_content, [INVOICE_EXTRACTION_PROMPT, file_part]
                )
            except Exception as exc:
                return {"filename": file_name, "success": False, "error": str(exc)}
        except Exception as exc:
            return {"filename": file_name, "success": False, "error": str(exc)}

    raw_text = _extract_text_from_genai_response(response)
    parsed, warnings, parse_err = parse_and_normalize_raw(raw_text)

    if parsed is None and GEMINI_INVOICE_JSON_RETRIES > 0:
        # Single bounded repair retry: re-ask the model to return valid JSON only.
        repair_prompt = (
            "Your previous response was not valid JSON. Re-emit the SAME extracted data as STRICT JSON\n"
            "matching the schema described earlier. No markdown, no commentary.\n\n"
            f"Previous response (truncated):\n{raw_text[:1500]}"
        )
        try:
            repaired = await asyncio.to_thread(
                model.generate_content,
                [repair_prompt, file_part],
                generation_config=generation_config,
            )
            repaired_text = _extract_text_from_genai_response(repaired)
            if repaired_text.strip():
                raw_text = repaired_text
            parsed, warnings, parse_err = parse_and_normalize_raw(raw_text)
        except Exception as exc:
            warnings = list(warnings) + [f"repair_retry_failed: {exc}"]

    if parsed is None:
        return {
            "filename": file_name,
            "success": False,
            "error": (parse_err or "model_output_not_json"),
            "raw_data": raw_text,
            "parsed": None,
            "warnings": warnings,
            "parse_error": parse_err,
        }

    canon = json.dumps(parsed, ensure_ascii=False, default=str)
    result: dict[str, Any] = {
        "filename": file_name,
        "success": True,
        "raw_data": canon,
        "parsed": parsed,
        "warnings": warnings,
    }
    return result


@app.post("/api/upload-invoice")
async def upload_invoice(
    invoice_pdf: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ["admin", "logistics", "user"]:
        raise HTTPException(status_code=403, detail="Authorized operations access required")
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=503, detail="AI extraction is not configured.")
    enforce_upload_rate_limit(f"{current_user.id}:upload_invoice")
    require_pdf_files(invoice_pdf)

    # Read all uploads sequentially (UploadFile streams must be consumed in order),
    # validate sizes, then split into 'cached' (hash hits) and 'fresh' (need Gemini).
    file_entries: list[dict] = []
    for file in invoice_pdf:
        content = await file.read()
        if len(content) > MAX_UPLOAD_FILE_SIZE_BYTES:
            raise HTTPException(status_code=400, detail=f"{file.filename} exceeds the maximum allowed size.")
        file_entries.append({
            "filename": file.filename or "invoice.pdf",
            "content": content,
            "sha256": compute_sha256(content),
        })

    # Idempotent cache: build placeholder results in original order, and only fan
    # out Gemini calls for entries we have not seen before.
    results: list[Optional[dict]] = [None] * len(file_entries)
    fresh_indices: list[int] = []
    for idx, entry in enumerate(file_entries):
        cached = find_cached_uploaded_document(db, entry["sha256"], "invoice")
        if cached and cached.status == "extracted":
            try:
                parsed_obj = json.loads(cached.parsed_json) if cached.parsed_json else None
            except Exception:
                parsed_obj = None
            try:
                warnings_obj = json.loads(cached.warnings_json) if cached.warnings_json else []
            except Exception:
                warnings_obj = []
            cache_ok = parsed_obj is not None
            raw_out = cached.raw_data or ""
            if cache_ok:
                try:
                    raw_out = json.dumps(parsed_obj, ensure_ascii=False, default=str)
                except Exception:
                    pass
            blob: dict[str, Any] = {
                "filename": entry["filename"],
                "success": cache_ok,
                "raw_data": raw_out,
                "parsed": parsed_obj,
                "warnings": list(warnings_obj) + ["duplicate_upload_cache_hit"],
                "duplicate": True,
                "cached_uploaded_at": cached.uploaded_at.isoformat() if cached.uploaded_at else None,
                "sha256": entry["sha256"],
            }
            if not cache_ok:
                blob["parse_error"] = cached.parse_error or "cached_row_missing_parsed_invoice"
                blob.setdefault("error", blob["parse_error"])
            results[idx] = blob
        else:
            fresh_indices.append(idx)

    if fresh_indices:
        model = genai.GenerativeModel(GEMINI_INVOICE_MODEL)
        semaphore = asyncio.Semaphore(GEMINI_INVOICE_MAX_CONCURRENCY)
        tasks = [
            _gemini_extract_one_invoice(model, file_entries[i]["filename"], file_entries[i]["content"], semaphore)
            for i in fresh_indices
        ]
        fresh_results = await asyncio.gather(*tasks)
        for slot, result in zip(fresh_indices, fresh_results):
            entry = file_entries[slot]
            result.setdefault("duplicate", False)
            result["sha256"] = entry["sha256"]
            results[slot] = result
            # Persist to cache only when extraction call itself succeeded.
            if result.get("success"):
                persist_uploaded_document(
                    db,
                    sha256_hex=entry["sha256"],
                    kind="invoice",
                    filename=entry["filename"],
                    byte_size=len(entry["content"]),
                    uploaded_by=current_user.username,
                    raw_data=result.get("raw_data"),
                    parsed=result.get("parsed"),
                    warnings=result.get("warnings") or [],
                    parse_error=result.get("parse_error"),
                )

    return {"success": True, "results": results}
PO_EXTRACTION_PROMPT = """
You are an expert Purchase Order data extraction assistant. Extract details from this PO and
return STRICT JSON only (no markdown, no commentary, no backticks).

CRITICAL INSTRUCTION FOR ITEMS:
  - For 'desc' you MUST capture the ENTIRE multi-line description for each line.
  - Use plain numbers (no currency symbols, no thousands separators).

Return JSON matching this structure:
{
  "poNo": "PO Number",
  "items": [
    {"desc": "ENTIRE paragraph of the ordered goods description exactly as written", "qty": 100, "uom": "MT"}
  ]
}
"""


async def _gemini_extract_one_po(model, file_name: str, content: bytes, semaphore: asyncio.Semaphore) -> dict:
    file_part = {"mime_type": "application/pdf", "data": content}
    generation_config = {"response_mime_type": "application/json"}
    async with semaphore:
        try:
            response = await asyncio.to_thread(
                model.generate_content, [PO_EXTRACTION_PROMPT, file_part], generation_config=generation_config
            )
        except TypeError:
            try:
                response = await asyncio.to_thread(model.generate_content, [PO_EXTRACTION_PROMPT, file_part])
            except Exception as exc:
                return {"filename": file_name, "success": False, "error": str(exc)}
        except Exception as exc:
            return {"filename": file_name, "success": False, "error": str(exc)}
    po_text = _extract_text_from_genai_response(response)
    return {"filename": file_name, "success": True, "raw_data": po_text}


@app.post("/api/upload-po")
async def upload_po(
    po_pdf: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=503, detail="AI extraction is not configured.")
    enforce_upload_rate_limit(f"{current_user.id}:upload_po")
    require_pdf_files(po_pdf)

    file_entries: list[dict] = []
    for file in po_pdf:
        content = await file.read()
        if len(content) > MAX_UPLOAD_FILE_SIZE_BYTES:
            raise HTTPException(status_code=400, detail=f"{file.filename} exceeds the maximum allowed size.")
        file_entries.append({
            "filename": file.filename or "po.pdf",
            "content": content,
            "sha256": compute_sha256(content),
        })

    results: list[Optional[dict]] = [None] * len(file_entries)
    fresh_indices: list[int] = []
    for idx, entry in enumerate(file_entries):
        cached = find_cached_uploaded_document(db, entry["sha256"], "po")
        if cached and cached.status == "extracted":
            results[idx] = {
                "filename": entry["filename"],
                "success": True,
                "raw_data": cached.raw_data or "",
                "duplicate": True,
                "cached_uploaded_at": cached.uploaded_at.isoformat() if cached.uploaded_at else None,
                "sha256": entry["sha256"],
            }
        else:
            fresh_indices.append(idx)

    if fresh_indices:
        model = genai.GenerativeModel(GEMINI_PO_MODEL)
        semaphore = asyncio.Semaphore(GEMINI_INVOICE_MAX_CONCURRENCY)
        tasks = [
            _gemini_extract_one_po(model, file_entries[i]["filename"], file_entries[i]["content"], semaphore)
            for i in fresh_indices
        ]
        fresh_results = await asyncio.gather(*tasks)
        for slot, result in zip(fresh_indices, fresh_results):
            entry = file_entries[slot]
            result.setdefault("duplicate", False)
            result["sha256"] = entry["sha256"]
            results[slot] = result
            if result.get("success"):
                persist_uploaded_document(
                    db,
                    sha256_hex=entry["sha256"],
                    kind="po",
                    filename=entry["filename"],
                    byte_size=len(entry["content"]),
                    uploaded_by=current_user.username,
                    raw_data=result.get("raw_data"),
                    parsed=None,
                    warnings=[],
                    parse_error=None,
                )

    return {"success": True, "results": results}

# --- PWA assets (served from site root for manifest / service worker scope) ---
@app.get("/manifest.webmanifest")
def serve_pwa_manifest():
    return FileResponse(
        "public/manifest.webmanifest",
        media_type="application/manifest+json",
    )


@app.get("/sw.js")
def serve_service_worker():
    return FileResponse(
        "public/sw.js",
        media_type="application/javascript",
        headers={"Cache-Control": "no-cache"},
    )


@app.get("/offline.html")
def serve_offline_page():
    return FileResponse("public/offline.html")


@app.get("/icon-192.png")
def serve_icon_192():
    return FileResponse("public/icon-192.png", media_type="image/png")


@app.get("/icon-512.png")
def serve_icon_512():
    return FileResponse("public/icon-512.png", media_type="image/png")


# --- Static File Routing (CRITICAL: MUST BE THE ABSOLUTE LAST LINES OF THE FILE) ---
app.mount("/static", StaticFiles(directory="public"), name="static")

@app.get("/")
def serve_frontend():
    return FileResponse("public/index.html")

if __name__ == "__main__":
    import uvicorn
    # Host / port are env-driven so the same binary works on a dev laptop, a LAN
    # box, or a container without code changes. See .env.example.
    host = os.getenv("HOST", "127.0.0.1").strip() or "127.0.0.1"
    port = int(os.getenv("PORT", "3000"))
    # Import string avoids __main__ vs main circular imports from routers/*.py.
    uvicorn.run("main:app", host=host, port=port)